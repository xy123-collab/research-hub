import io
import hashlib
import json
from collections import Counter
from decimal import Decimal, InvalidOperation
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from ..core.db import get_db
from ..core.storage import storage
from ..core.permissions import (get_current_user, is_dataset_member, is_dataset_admin,
                                count_dataset_admins, get_settings)
from ..core.audit import write_audit, record_contribution
from ..models.user import User
from ..models.dataset import Dataset, Variable
from ..models.version import DataVersion
from ..models.correction import Bug, CorrectionReview, CorrectionFinal
from ..models.curation import BugItem, VersionExtra, DatasetDataConfig
from ..schemas.models import (
    BugIn, UnstructuredBugIn, ReviewIn, FinalizeIn, PartialFinalizeIn,
)
from ..core.time import china_iso

router = APIRouter(tags=["bugs"])


def _uid_var(db, dataset_id) -> str | None:
    c = db.get(DatasetDataConfig, dataset_id)
    return c.unique_id_var if c else None


def _ds(db, slug):
    d = db.query(Dataset).filter_by(slug=slug, is_deleted=False).first()
    if not d:
        raise HTTPException(404, "数据集不存在")
    return d


def _latest_raw_version(db: Session, dataset_id: int) -> DataVersion | None:
    rows = db.query(DataVersion).filter_by(dataset_id=dataset_id).order_by(
        DataVersion.id.desc()).all()
    for version in rows:
        if not version.data_file_path:
            continue
        extra = db.get(VersionExtra, version.id)
        if (extra.data_kind if extra else "raw") == "raw":
            return version
    return None


def _canonical_uid(value, numeric: bool) -> str:
    if value is None:
        return ""
    try:
        import pandas as pd
        if pd.isna(value):
            return ""
    except Exception:
        pass
    text = str(value).strip()
    if not numeric:
        return text
    try:
        return format(Decimal(text).normalize(), "f")
    except (InvalidOperation, ValueError):
        return text


def _uid_context(db: Session, dataset_id: int, requested_var: str | None = None) -> dict:
    """读取本次选用的 ID 列，并建立规范化计数索引。"""
    cfg = db.get(DatasetDataConfig, dataset_id)
    recommended_var = cfg.unique_id_var if cfg else None
    if not recommended_var:
        raise HTTPException(400, "管理员尚未设置唯一 ID 变量，请先到版本库的「数据处理设置」配置")
    uid_var = str(requested_var or recommended_var).strip()
    variable = db.query(Variable).filter_by(
        dataset_id=dataset_id, var_name=uid_var, enabled=True).first()
    if not variable:
        raise HTTPException(400, f"ID 变量「{uid_var}」不在当前变量清单")
    version = _latest_raw_version(db, dataset_id)
    if not version:
        raise HTTPException(400, "还没有可用于核对唯一 ID 的原始数据版本，请管理员先发布原始数据")
    try:
        from ..services.introspect import read_table_column
        import pandas as pd
        series = read_table_column(version.data_file_path, uid_var)
        numeric = bool(pd.api.types.is_numeric_dtype(series))
        counts = Counter(_canonical_uid(value, numeric) for value in series)
    except Exception as exc:
        raise HTTPException(
            400, f"无法读取最新原始版本「{version.version_id}」的唯一 ID：{exc}。"
                 "请检查文件是否仍在 COS、格式是否正确及唯一 ID 配置") from exc
    return {"unique_id_var": uid_var, "recommended_uid_var": recommended_var,
            "uses_alternative_id": uid_var != recommended_var,
            "version": version, "numeric": numeric, "counts": counts}


def _check_uid(context: dict, value) -> dict:
    raw = str(value or "").strip()
    if not raw:
        raise HTTPException(400, f"请填写{context['unique_id_var']}（唯一 ID）")
    normalized = _canonical_uid(raw, context["numeric"])
    count = int(context["counts"].get(normalized, 0))
    if count > 1:
        raise HTTPException(
            400, f"{context['unique_id_var']}={raw} 在最新原始版本中匹配到 {count} 条；"
                 "唯一 ID 不唯一，请管理员先修正数据或配置")
    return {"value": raw, "normalized": normalized, "count": count, "exists": count == 1}


def _match_text(value) -> str:
    return str(value if value is not None else "").strip()


def _bug_signature(context: dict, uid_value, var_name, current_value, suggested_value) -> tuple:
    """重复勘误按 ID 变量、ID 值、修改变量、当前值、建议值匹配。"""
    return (
        context["unique_id_var"],
        _canonical_uid(uid_value, context["numeric"]),
        _match_text(var_name),
        _match_text(current_value),
        _match_text(suggested_value),
    )


def _duplicate_index(db: Session, dataset_id: int, context: dict,
                     exclude_bug_id: int | None = None,
                     exclude_item_id: int | None = None) -> dict[tuple, dict]:
    """覆盖所有状态的历史勘误；部分采纳后同时索引原始内容和管理员修改内容。"""
    index: dict[tuple, dict] = {}
    contexts = {context["unique_id_var"]: context}

    def context_for(uid_var: str | None) -> dict:
        selected = (uid_var or context["recommended_uid_var"]).strip()
        if selected not in contexts:
            try:
                contexts[selected] = _uid_context(db, dataset_id, selected)
            except HTTPException:
                # 历史勘误使用的 ID 变量可能已在后续版本中停用；仍保留文本口径
                # 参与历史查重，但不能因此阻断当前版本的新提交。
                contexts[selected] = {
                    "unique_id_var": selected,
                    "recommended_uid_var": context["recommended_uid_var"],
                    "uses_alternative_id": selected != context["recommended_uid_var"],
                    "numeric": False,
                }
        return contexts[selected]

    item_bug_ids = set()
    rows = db.query(BugItem, Bug).join(Bug, Bug.id == BugItem.bug_id).filter(
        BugItem.dataset_id == dataset_id).all()
    for item, bug in rows:
        item_bug_ids.add(bug.id)
        if bug.id == exclude_bug_id or item.id == exclude_item_id:
            continue
        meta = {"bug_id": bug.id, "item_id": item.id,
                "status": item.status or bug.status or "pending"}
        item_context = context_for(item.uid_var)
        current = _bug_signature(
            item_context, item.uid_value, item.var_name,
            item.current_value, item.suggested_value,
        )
        index.setdefault(current, meta)
        if item.original_uid_value is not None:
            original = _bug_signature(
                item_context, item.original_uid_value, item.original_var_name,
                item.original_current_value, item.original_suggested_value,
            )
            index.setdefault(original, meta)

    # 兼容尚未生成 BugItem 的旧勘误。
    variables = {v.id: v.var_name for v in db.query(Variable).filter_by(
        dataset_id=dataset_id).all()}
    for bug in db.query(Bug).filter_by(dataset_id=dataset_id).all():
        if bug.id in item_bug_ids or bug.id == exclude_bug_id:
            continue
        if bug.bug_type == "unstructured":
            continue
        signature = _bug_signature(
            context, bug.officer_id or bug.term_id,
            variables.get(bug.variable_id, ""),
            bug.current_value, bug.suggested_value,
        )
        index.setdefault(signature, {
            "bug_id": bug.id, "item_id": None, "status": bug.status or "pending",
        })
    return index


def _duplicate_message(duplicate: dict) -> str:
    labels = {
        "pending": "未确认", "accepted": "已采纳未修改",
        "fixed": "已采纳已修改", "rejected": "未采纳",
    }
    status = labels.get(duplicate.get("status"), duplicate.get("status") or "未知")
    return f"与历史勘误 #{duplicate['bug_id']} 重复（当前状态：{status}）"


def _effective_bug_status(items: list[BugItem], fallback: str) -> str:
    if not items:
        return fallback
    states = {item.status for item in items}
    if "pending" in states:
        return "pending"
    if "accepted" in states:
        return "accepted"
    if "fixed" in states:
        return "fixed"
    return "rejected"


def _sync_bug_status(db: Session, bug: Bug, fixed_version_id: int | None = None) -> str:
    items = db.query(BugItem).filter_by(bug_id=bug.id).all()
    bug.status = _effective_bug_status(items, bug.status)
    if bug.status == "fixed" and fixed_version_id:
        bug.fixed_in_version_id = fixed_version_id
    return bug.status


@router.get("/datasets/{slug}/bugs")
def list_bugs(slug: str, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    d = _ds(db, slug)
    bugs = db.query(Bug).filter_by(dataset_id=d.id).order_by(Bug.id.desc()).all()
    out = []
    for b in bugs:
        items = db.query(BugItem).filter_by(bug_id=b.id).order_by(BugItem.seq).all()
        reporter = db.get(User, b.reporter_id)
        reviewer_count = db.query(CorrectionReview.reviewer_id).filter_by(
            target_type="bug", target_id=b.id, reviewer_type="member").filter(
            CorrectionReview.reviewer_id.is_not(None)).distinct().count()
        out.append({
            "id": b.id, "officer_id": b.officer_id, "term_id": b.term_id,
            "current_value": b.current_value, "suggested_value": b.suggested_value,
            "description_zh": b.description_zh,
            "correction_type": (
                "unstructured" if b.bug_type == "unstructured" else "structured"
            ),
            "status": _effective_bug_status(items, b.status),
            "reporter_id": b.reporter_id,
            "reporter_name": reporter.display_name if reporter else "已注销用户",
            "created_at": china_iso(b.created_at),
            "reviewer_count": reviewer_count,
            "has_new_officer": any(bool(it.is_new_officer) for it in items),
            "has_manual_only": (
                b.bug_type == "unstructured"
                or any(bool(it.is_new_officer) for it in items)
            ),
            "can_delete": (
                b.reporter_id == user.id and b.status == "pending"
                and all(it.status == "pending" for it in items)
            ),
        })
    return out


@router.get("/datasets/{slug}/bugs/id-check")
def check_bug_uid(slug: str, value: str, db: Session = Depends(get_db),
                  id_var: str | None = None,
                  user: User = Depends(get_current_user)):
    d = _ds(db, slug)
    if not is_dataset_member(db, d.id, user):
        raise HTTPException(403, "需为数据集成员")
    context = _uid_context(db, d.id, id_var)
    result = _check_uid(context, value)
    return {"unique_id_var": context["unique_id_var"],
            "recommended_unique_id_var": context["recommended_uid_var"],
            "uses_alternative_id": context["uses_alternative_id"],
            "source_version": context["version"].version_id, **result}


@router.post("/datasets/{slug}/bugs")
def submit_bug(slug: str, body: dict, db: Session = Depends(get_db),
               user: User = Depends(get_current_user)):
    d = _ds(db, slug)
    if not is_dataset_member(db, d.id, user):
        raise HTTPException(403, "非成员不能提交勘误，请先申请加入处理")
    if get_settings(db, d.id).is_closed:
        raise HTTPException(400, "数据集已关闭，不再接受新勘误")
    try:
        body = BugIn.model_validate(body)
    except Exception as exc:
        raise HTTPException(422, "请完整填写勘误说明和证据") from exc
    context = _uid_context(db, d.id, body.uid_var)
    if context["uses_alternative_id"] and not body.confirm_alternative_id:
        raise HTTPException(
            409, f"本次选择的 ID 变量「{context['unique_id_var']}」不是管理员推荐的"
                 f"「{context['recommended_uid_var']}」。请确认已理解定位口径后再提交")
    uid_value = body.officer_id or body.term_id or ""
    check = _check_uid(context, uid_value)
    if not check["exists"] and not body.confirm_new_officer:
        raise HTTPException(
            409, f"最新原始版本中没有 {context['unique_id_var']}={check['value']}。"
                 "请确认 ID 是否填错；若确为新增主体或新增记录，请勾选确认后再提交")
    var = db.get(Variable, body.variable_id) if body.variable_id else None
    if not var or var.dataset_id != d.id or not var.enabled:
        raise HTTPException(400, "请选择本数据集当前有效的勘误变量")
    if var.var_name in {
            context["unique_id_var"], context["recommended_uid_var"]}:
        raise HTTPException(400, "管理员推荐或本次选用的 ID 变量本身不能作为勘误修改对象")
    duplicate = _duplicate_index(db, d.id, context).get(_bug_signature(
        context, check["value"], var.var_name,
        body.current_value, body.suggested_value,
    ))
    if duplicate:
        raise HTTPException(409, "重复勘误，无法提交：" + _duplicate_message(duplicate))
    data = body.model_dump(exclude={
        "uid_var", "confirm_new_officer", "confirm_alternative_id",
    })
    b = Bug(dataset_id=d.id, reporter_id=user.id, status="pending",
            created_at=datetime.utcnow(), **data)
    db.add(b); db.flush()
    # 单条勘误也建一个子项，统一按子项打分/终审/应用
    db.add(BugItem(bug_id=b.id, dataset_id=d.id, seq=1,
                   uid_var=context["unique_id_var"], uid_value=check["value"],
                   var_name=var.var_name,
                   current_value=body.current_value, suggested_value=body.suggested_value,
                   reason=body.description_zh, evidence=body.evidence,
                   is_new_officer=not check["exists"], status="pending"))
    write_audit(db, user.id, "bug.submit", "bug", b.id,
                {"is_new_officer": not check["exists"],
                 "uid_var": context["unique_id_var"],
                 "uses_alternative_id": context["uses_alternative_id"],
                 "checked_version": context["version"].version_id})
    db.commit()
    return {"id": b.id}


@router.post("/datasets/{slug}/bugs/unstructured")
def submit_unstructured_bug(
        slug: str, body: UnstructuredBugIn, db: Session = Depends(get_db),
        user: User = Depends(get_current_user)):
    """提交无法按单元格描述的自然语言勘误；只进入人工审核和人工修改流程。"""
    d = _ds(db, slug)
    if not is_dataset_member(db, d.id, user):
        raise HTTPException(403, "非成员不能提交勘误，请先申请加入处理")
    if get_settings(db, d.id).is_closed:
        raise HTTPException(400, "数据集已关闭，不再接受新勘误")
    b = Bug(
        dataset_id=d.id, reporter_id=user.id, status="pending",
        bug_type="unstructured", description_zh=body.issue,
        suggested_value=body.suggestion, evidence=body.evidence,
        created_at=datetime.utcnow(),
    )
    db.add(b)
    db.flush()
    write_audit(db, user.id, "bug.submit.unstructured", "bug", b.id, {
        "manual_only": True,
    })
    db.commit()
    return {"id": b.id, "manual_only": True}


def _var_name(db, variable_id):
    v = db.get(Variable, variable_id)
    return v.var_name if v else ""


# ============ 批量勘误：模板下载 + Excel/CSV 导入 ============
LOCATOR_VAR_COL = "定位唯一id"
LOCATOR_VALUE_COL = "定位id的值"
BATCH_COLS = [
    LOCATOR_VAR_COL, LOCATOR_VALUE_COL,
    "变量名", "当前值", "建议值", "说明", "证据",
]
BATCH_REQUIRED = [LOCATOR_VALUE_COL, "变量名", "说明", "证据"]
LEGACY_LOCATOR_VAR_COL = "ID变量名"
LEGACY_LOCATOR_VALUE_COL = "唯一ID值"
COMMON_BATCH_COLS = ["变量名", "当前值", "建议值", "说明", "证据"]


@router.get("/datasets/{slug}/bug-template")
def bug_template(slug: str, db: Session = Depends(get_db),
                 user: User = Depends(get_current_user)):
    """下载批量勘误 Excel 模板：含列头、规则注释与变量清单。"""
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    d = _ds(db, slug)
    uidv = _uid_var(db, d.id)
    wb = Workbook()
    ws = wb.active; ws.title = "勘误"
    ws.append(BATCH_COLS)
    notes = {
        LOCATOR_VAR_COL: f"可留空，默认使用管理员推荐的「{uidv or '尚未设置'}」。"
                         "如填写其他定位 ID，提交前必须额外确认；只要最新数据中恰好匹配一行，仍可安全一键应用。",
        LOCATOR_VALUE_COL: "填写本行用于定位记录的 ID 值。",
        "变量名": "要修改的变量名（须与数据集变量一致）。",
        "当前值": "该单元格现在的值。",
        "建议值": "建议改成的值。",
        "说明": "为什么需要修改（必填）。",
        "证据": "证据来源、链接或出处（必填；附件可在单条勘误中另行上传）。",
    }
    for i, col in enumerate(BATCH_COLS, start=1):
        ws.cell(row=1, column=i).comment = Comment(notes[col], "系统")
    # 变量清单页，方便对照
    ws2 = wb.create_sheet("变量清单")
    ws2.append(["变量名", "标签"])
    for v in db.query(Variable).filter_by(dataset_id=d.id, enabled=True).all():
        ws2.append([v.var_name, v.label_zh or ""])
    ws3 = wb.create_sheet("填写说明")
    for line in [
        "批量勘误填写说明：",
        "1. 每一行是一个独立勘误；提交后会在勘误列表中分别显示、分别打分和分别终审。",
        "2. 上传文件后先逐行校验；有问题的行会显示具体原因，可删除问题行后再提交其余行。",
        f"3. 管理员推荐定位 ID：{uidv or '（管理员尚未设置，请先在数据集设置里指定唯一ID）'}。",
        "4. 第一列「定位唯一id」可留空；如不用管理员推荐 ID，须填写变量名并在上传页额外确认。",
        "5. ID 变量本身不允许被修改；要修改的变量名须与「变量清单」页一致。",
        "6. 「说明」和「证据」是两个独立必填列，旧版合并列模板不再接受。",
        "7. 系统会用「定位唯一id、定位id的值、变量名、当前值、建议值」与全部历史勘误查重。",
        "8. 支持 .xlsx / .csv，列顺序：" + "、".join(BATCH_COLS) + "。",
    ]:
        ws3.append([line])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 f'attachment; filename="{d.slug}_bug_template.xlsx"'})


def _validate_batch_rows(db: Session, dataset: Dataset, rows: list[dict],
                         included_row_numbers: set[int] | None = None
                         ) -> tuple[dict, list[dict]]:
    if not rows:
        raise HTTPException(400, "未解析到有效数据行，请用最新模板填写")
    header = set(rows[0].keys())
    if "说明与证据" in header:
        raise HTTPException(400, "这是旧版批量模板。请重新下载模板，将「说明」和「证据」分列填写")
    missing_cols = [col for col in COMMON_BATCH_COLS if col not in header]
    if LOCATOR_VALUE_COL not in header and LEGACY_LOCATOR_VALUE_COL not in header:
        missing_cols.insert(0, LOCATOR_VALUE_COL)
    if missing_cols:
        raise HTTPException(400, "批量模板缺少列：" + "、".join(missing_cols))
    context = _uid_context(db, dataset.id)
    contexts = {context["unique_id_var"]: context}
    variables = {v.var_name for v in db.query(Variable).filter_by(
        dataset_id=dataset.id, enabled=True).all()}
    historical = _duplicate_index(db, dataset.id, context)
    seen_batch: dict[tuple, int] = {}
    checked = []
    for row_no, row in enumerate(rows, start=2):
        if included_row_numbers is not None and row_no not in included_row_numbers:
            continue
        cleaned = {
            LOCATOR_VAR_COL: str(
                row.get(LOCATOR_VAR_COL, row.get(LEGACY_LOCATOR_VAR_COL, "")) or ""
            ).strip(),
            LOCATOR_VALUE_COL: str(
                row.get(LOCATOR_VALUE_COL, row.get(LEGACY_LOCATOR_VALUE_COL, "")) or ""
            ).strip(),
            **{
                key: str(row.get(key, "") if row.get(key) is not None else "").strip()
                for key in COMMON_BATCH_COLS
            },
        }
        problems = []
        missing = [col for col in BATCH_REQUIRED if not cleaned[col]]
        if missing:
            problems.append("缺少必填项：" + "、".join(missing))
        if cleaned["变量名"] not in variables:
            problems.append(f"变量「{cleaned['变量名']}」不在当前变量清单")
        uid_check = None
        selected_uid_var = cleaned[LOCATOR_VAR_COL] or context["recommended_uid_var"]
        row_context = contexts.get(selected_uid_var)
        if row_context is None:
            try:
                row_context = _uid_context(db, dataset.id, selected_uid_var)
                contexts[selected_uid_var] = row_context
            except HTTPException as exc:
                problems.append(str(exc.detail))
        if cleaned["变量名"] in {
                context["recommended_uid_var"], selected_uid_var}:
            problems.append("不能修改管理员推荐或本次选用的 ID 变量本身")
        if cleaned[LOCATOR_VALUE_COL]:
            if row_context is not None:
                try:
                    uid_check = _check_uid(row_context, cleaned[LOCATOR_VALUE_COL])
                except HTTPException as exc:
                    problems.append(str(exc.detail))
        if uid_check and row_context and cleaned["变量名"] in variables:
            signature = _bug_signature(
                row_context, uid_check["value"], cleaned["变量名"],
                cleaned["当前值"], cleaned["建议值"],
            )
            duplicate = historical.get(signature)
            if duplicate:
                problems.append("重复勘误：" + _duplicate_message(duplicate))
            elif not problems and signature in seen_batch:
                problems.append(f"与本批第 {seen_batch[signature]} 行重复")
            elif not problems:
                seen_batch[signature] = row_no
        checked.append({
            "row_no": row_no, "row": cleaned, "uid": uid_check,
            "uid_context": row_context,
            "valid": not problems, "problems": problems,
        })
    return context, checked


def _public_batch_item(item: dict) -> dict:
    uid = item.get("uid")
    return {
        "row_no": item["row_no"],
        **item["row"],
        "valid": item["valid"],
        "problems": item["problems"],
        "uid_exists": bool(uid and uid["exists"]),
        "is_new_officer": bool(uid and not uid["exists"]),
        "uses_alternative_id": bool(
            item.get("uid_context")
            and item["uid_context"]["uses_alternative_id"]
        ),
    }


@router.post("/datasets/{slug}/bugs/batch/validate")
def validate_bug_batch(slug: str, file: UploadFile = File(...),
                       db: Session = Depends(get_db),
                       user: User = Depends(get_current_user)):
    d = _ds(db, slug)
    if not is_dataset_member(db, d.id, user):
        raise HTTPException(403, "非成员不能提交勘误")
    rows = _parse_batch(file.file.read(), file.filename)
    context, checked = _validate_batch_rows(db, d, rows)
    missing = []
    seen = set()
    for item in checked:
        if (item["valid"] and item["uid"] and not item["uid"]["exists"]
                and item["uid"]["value"] not in seen):
            seen.add(item["uid"]["value"])
            missing.append(item["uid"]["value"])
    return {"rows": len(checked),
            "valid_rows": sum(1 for item in checked if item["valid"]),
            "invalid_rows": sum(1 for item in checked if not item["valid"]),
            "problem_rows": sum(
                1 for item in checked
                if (not item["valid"]
                    or bool(item["uid"] and not item["uid"]["exists"])
                    or bool(item["uid_context"]
                            and item["uid_context"]["uses_alternative_id"]))
            ),
            "items": [_public_batch_item(item) for item in checked],
            "unique_id_var": context["unique_id_var"],
            "source_version": context["version"].version_id,
            "missing_uid_values": missing,
            "alternative_id_rows": [
                item["row_no"] for item in checked
                if item["valid"] and item["uid_context"]["uses_alternative_id"]
            ]}


@router.post("/datasets/{slug}/bugs/batch")
def submit_bug_batch(slug: str, file: UploadFile = File(...), title: str = Form(""),
                     confirmed_new_officer_ids: str = Form("[]"),
                     confirmed_alternative_id_rows: str = Form("[]"),
                     included_row_numbers: str = Form(""),
                     db: Session = Depends(get_db),
                     user: User = Depends(get_current_user)):
    """批量导入：每个通过校验且被保留的文件行生成一条独立勘误。"""
    d = _ds(db, slug)
    if not is_dataset_member(db, d.id, user):
        raise HTTPException(403, "非成员不能提交勘误")
    if get_settings(db, d.id).is_closed:
        raise HTTPException(400, "数据集已关闭，不再接受新勘误")
    raw = file.file.read()
    rows = _parse_batch(raw, file.filename)
    try:
        confirmed = {str(x).strip() for x in json.loads(confirmed_new_officer_ids or "[]")}
    except Exception:
        raise HTTPException(400, "新增主体/记录确认参数格式错误，请重新校验批量文件")
    try:
        confirmed_alternative_rows = {
            int(x) for x in json.loads(confirmed_alternative_id_rows or "[]")
        }
    except Exception:
        raise HTTPException(400, "非推荐 ID 确认参数格式错误，请重新校验批量文件")
    known_rows = set(range(2, len(rows) + 2))
    try:
        requested = (
            {int(x) for x in json.loads(included_row_numbers)}
            if included_row_numbers.strip() else known_rows
        )
    except Exception:
        raise HTTPException(400, "保留行参数格式错误，请重新解析批量文件")
    if not requested:
        raise HTTPException(400, "没有可提交的勘误行，请至少保留一行")
    if not requested.issubset(known_rows):
        raise HTTPException(400, "批量文件内容已经变化，请重新解析并校验")
    context, selected = _validate_batch_rows(
        db, d, rows, included_row_numbers=requested)
    invalid = [item for item in selected if not item["valid"]]
    if invalid:
        details = "；".join(
            f"第 {item['row_no']} 行：" + "、".join(item["problems"])
            for item in invalid[:20]
        )
        raise HTTPException(409, "仍有问题行，删除后才能提交：" + details)
    unconfirmed = sorted({item["uid"]["value"] for item in selected
                          if not item["uid"]["exists"] and item["uid"]["value"] not in confirmed})
    if unconfirmed:
        raise HTTPException(
            409, "以下 ID 在最新原始版本中不存在，请确认它们是新增主体或新增记录后再提交："
                 + "、".join(unconfirmed[:20]))
    alternative_rows = {
        item["row_no"] for item in selected
        if item["uid_context"]["uses_alternative_id"]
    }
    unconfirmed_alternative = sorted(alternative_rows - confirmed_alternative_rows)
    if unconfirmed_alternative:
        raise HTTPException(
            409, "以下行没有使用管理员推荐的唯一 ID，请确认定位口径后再提交："
                 + "、".join(str(row_no) for row_no in unconfirmed_alternative[:20]))
    variables = {v.var_name: v for v in db.query(Variable).filter_by(
        dataset_id=d.id, enabled=True).all()}
    created_ids = []
    batch_label = ("batch:" + title.strip())[:60] if title.strip() else "batch"
    for item in selected:
        r = item["row"]
        variable = variables[r["变量名"]]
        b = Bug(
            dataset_id=d.id, reporter_id=user.id, status="pending",
            officer_id=item["uid"]["value"], variable_id=variable.id,
            current_value=r["当前值"], suggested_value=r["建议值"],
            bug_type=batch_label, description_zh=r["说明"], evidence=r["证据"],
            created_at=datetime.utcnow(),
        )
        db.add(b); db.flush()
        db.add(BugItem(bug_id=b.id, dataset_id=d.id, seq=1,
                       uid_var=item["uid_context"]["unique_id_var"],
                       uid_value=r[LOCATOR_VALUE_COL], var_name=r["变量名"],
                       current_value=r["当前值"], suggested_value=r["建议值"],
                       reason=r["说明"], evidence=r["证据"],
                       is_new_officer=not item["uid"]["exists"], status="pending"))
        created_ids.append(b.id)
        write_audit(db, user.id, "bug.submit.batch.item", "bug", b.id, {
            "source_row": item["row_no"],
            "is_new_officer": not item["uid"]["exists"],
            "uid_var": item["uid_context"]["unique_id_var"],
            "uses_alternative_id": item["uid_context"]["uses_alternative_id"],
            "checked_version": context["version"].version_id,
        })
    db.commit()
    return {"id": created_ids[0], "ids": created_ids, "items": len(created_ids)}


def _parse_batch(raw: bytes, filename: str) -> list[dict]:
    name = (filename or "").lower()
    rows = []
    if name.endswith(".csv") or (not name.endswith(".xlsx") and b"," in raw[:200]):
        import csv
        text = raw.decode("utf-8-sig", errors="replace")
        for r in csv.DictReader(io.StringIO(text)):
            if any((v or "").strip() for v in r.values()):
                rows.append(r)
    else:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb["勘误"] if "勘误" in wb.sheetnames else wb.worksheets[0]
        header = None
        for row in ws.iter_rows(values_only=True):
            if header is None:
                header = [str(c).strip() if c is not None else "" for c in row]
                continue
            if row is None or all(c is None for c in row):
                continue
            rows.append({header[i]: row[i] for i in range(min(len(header), len(row)))})
    return rows


@router.patch("/bugs/{bid}")
def edit_bug(bid: int, body: BugIn, db: Session = Depends(get_db),
             user: User = Depends(get_current_user)):
    b = db.get(Bug, bid)
    if not b:
        raise HTTPException(404, "勘误不存在")
    if b.reporter_id != user.id or b.status != "pending":
        raise HTTPException(403, "仅本人且未审核前可改")
    context = _uid_context(db, b.dataset_id, body.uid_var)
    if context["uses_alternative_id"] and not body.confirm_alternative_id:
        raise HTTPException(
            409, "本次没有使用管理员推荐的唯一 ID，请明确确认定位口径后再保存")
    check = _check_uid(context, body.officer_id or body.term_id)
    if not check["exists"] and not body.confirm_new_officer:
        raise HTTPException(409, "该 ID 不存在；若确为新增主体或新增记录，请勾选确认后再保存")
    var = db.get(Variable, body.variable_id) if body.variable_id else None
    if not var or var.dataset_id != b.dataset_id or not var.enabled:
        raise HTTPException(400, "请选择本数据集当前有效的勘误变量")
    item = db.query(BugItem).filter_by(bug_id=b.id).order_by(BugItem.seq).first()
    duplicate = _duplicate_index(
        db, b.dataset_id, context, exclude_bug_id=b.id,
        exclude_item_id=item.id if item else None,
    ).get(_bug_signature(
        context, check["value"], var.var_name,
        body.current_value, body.suggested_value,
    ))
    if duplicate:
        raise HTTPException(409, "重复勘误，无法保存：" + _duplicate_message(duplicate))
    for k, v in body.model_dump(
            exclude={"uid_var", "confirm_new_officer", "confirm_alternative_id"},
            exclude_none=True).items():
        setattr(b, k, v)
    if item:
        item.uid_var = context["unique_id_var"]
        item.uid_value = check["value"]
        item.var_name = var.var_name
        item.current_value = body.current_value
        item.suggested_value = body.suggested_value
        item.reason = body.description_zh
        item.evidence = body.evidence
        item.is_new_officer = not check["exists"]
    db.commit()
    return {"ok": True}


@router.delete("/bugs/{bid}")
def del_bug(bid: int, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    b = db.get(Bug, bid)
    if not b:
        raise HTTPException(404, "勘误不存在")
    items = db.query(BugItem).filter_by(bug_id=bid).all()
    if (b.reporter_id != user.id or b.status != "pending"
            or any(item.status != "pending" for item in items)):
        raise HTTPException(403, "仅本人且未审核前可删")
    from ..models.correction import BugAttachment
    attachment_paths = [
        row.file_path for row in db.query(BugAttachment).filter_by(bug_id=bid).all()
        if row.file_path
    ]
    item_ids = [item.id for item in items]
    db.query(CorrectionReview).filter_by(target_type="bug", target_id=bid).delete(
        synchronize_session=False)
    if item_ids:
        db.query(CorrectionReview).filter_by(target_type="bug_item").filter(
            CorrectionReview.target_id.in_(item_ids)).delete(synchronize_session=False)
        db.query(CorrectionFinal).filter_by(target_type="bug_item").filter(
            CorrectionFinal.target_id.in_(item_ids)).delete(synchronize_session=False)
    db.query(CorrectionFinal).filter_by(target_type="bug", target_id=bid).delete(
        synchronize_session=False)
    db.query(BugAttachment).filter_by(bug_id=bid).delete(synchronize_session=False)
    db.query(BugItem).filter_by(bug_id=bid).delete(synchronize_session=False)
    db.query(Bug).filter_by(id=bid).delete(synchronize_session=False)
    write_audit(db, user.id, "bug.delete", "bug", bid)
    db.commit()
    for file_path in attachment_paths:
        try:
            storage.delete(file_path)
        except Exception:
            pass
    return {"ok": True}


@router.get("/bugs/{bid}/reviews")
def get_reviews(bid: int, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    rs = db.query(CorrectionReview).filter_by(target_type="bug", target_id=bid).all()
    fin = db.query(CorrectionFinal).filter_by(target_type="bug", target_id=bid).first()
    return {"reviews": [{"reviewer_type": r.reviewer_type, "reviewer_id": r.reviewer_id,
                         "score": r.acceptability_score, "comment": r.comment} for r in rs],
            "final": ({"adopt_level": fin.adopt_level, "final_score": fin.final_score,
                       "comment": fin.comment} if fin else None)}


@router.post("/bugs/{bid}/reviews")
def review_bug(bid: int, body: ReviewIn, db: Session = Depends(get_db),
               user: User = Depends(get_current_user)):
    b = db.get(Bug, bid)
    if not b:
        raise HTTPException(404, "勘误不存在")
    if not is_dataset_member(db, b.dataset_id, user):
        raise HTTPException(403, "需为数据集成员")
    if _effective_bug_status(
            db.query(BugItem).filter_by(bug_id=bid).all(), b.status) != "pending":
        raise HTTPException(400, "该勘误已终审，不能继续评分")
    review = db.query(CorrectionReview).filter_by(
        target_type="bug", target_id=bid, reviewer_type="member",
        reviewer_id=user.id).first()
    if not review:
        review = CorrectionReview(target_type="bug", target_id=bid,
                                  reviewer_type="member", reviewer_id=user.id)
        db.add(review)
    review.acceptability_score = body.acceptability_score
    review.comment = body.comment
    db.commit()
    return {"ok": True}


def _parse_ai_review(response: str) -> tuple[float, str]:
    try:
        start, end = response.find("{"), response.rfind("}")
        payload = json.loads(response[start:end + 1])
        score = float(payload["score"])
        reason = str(payload["reason"]).strip()
        if not 0 <= score <= 10 or not reason:
            raise ValueError
        return score, reason[:1000]
    except Exception as exc:
        raise HTTPException(502, "AI 评分返回格式异常，未保存任何分数，请稍后重试") from exc


def _run_ai_review(prompt: str) -> tuple[float, str]:
    from ..core.ai_client import ai_client
    if not ai_client.enabled():
        raise HTTPException(400, "AI 勘误评分尚未启用，请联系平台管理员配置 AI")
    response = ai_client.complete(
        prompt,
        '你是数据质量评审助手。只输出 JSON：{"score":0到10的数字,"reason":"具体评分理由"}。'
        "必须同时比较当前值和建议值，并结合修改说明与证据；不得只给分不解释。",
    )
    return _parse_ai_review(response)


@router.post("/bugs/{bid}/ai-review")
def ai_review_bug(bid: int, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    b = db.get(Bug, bid)
    if not b:
        raise HTTPException(404, "勘误不存在")
    if not is_dataset_member(db, b.dataset_id, user):
        raise HTTPException(403, "需为数据集成员")
    if not (b.description_zh or "").strip() or not (b.evidence or "").strip():
        raise HTTPException(400, "该历史勘误缺少独立的说明或证据，补充完整后才能进行 AI 评分")
    prompt = (f"当前值：{b.current_value or ''}\n建议值：{b.suggested_value or ''}\n"
              f"修改说明：{b.description_zh}\n证据：{b.evidence}")
    score, reason = _run_ai_review(prompt)
    review = db.query(CorrectionReview).filter_by(
        target_type="bug", target_id=bid, reviewer_type="ai").first()
    if not review:
        review = CorrectionReview(target_type="bug", target_id=bid, reviewer_type="ai")
        db.add(review)
    review.acceptability_score = score
    review.comment = reason
    db.commit()
    return {"ai_score": score, "reason": reason}


@router.post("/bugs/{bid}/finalize")
def finalize_bug(bid: int, body: FinalizeIn, db: Session = Depends(get_db),
                 user: User = Depends(get_current_user)):
    b = db.get(Bug, bid)
    if not b:
        raise HTTPException(404, "勘误不存在")
    if not is_dataset_admin(db, b.dataset_id, user):
        raise HTTPException(403, "仅数据集管理员可审核勘误")
    if b.status != "pending":
        raise HTTPException(400, "该勘误已处理")
    if body.adopt_level == "partial":
        raise HTTPException(
            409, "部分采纳不能直接确认；请先修改唯一值、变量、当前值、建议值或理由，"
                 "再从部分采纳编辑页确认")
    # 八节 4：管理员本人提交的勘误，若还有其他管理员，应由另一名管理员审核
    self_review = (b.reporter_id == user.id)
    if self_review and count_dataset_admins(db, b.dataset_id) > 1:
        raise HTTPException(403, "这是你本人提交的勘误，请由另一名数据集管理员审核")
    db.add(CorrectionFinal(target_type="bug", target_id=bid, decided_by=user.id,
                           adopt_level=body.adopt_level, final_score=body.final_score,
                           comment=body.comment, decided_at=datetime.utcnow()))
    b.status = "accepted" if body.adopt_level != "reject" else "rejected"
    b.reviewed_by = user.id; b.reviewed_at = datetime.utcnow()
    child_status = "accepted" if body.adopt_level != "reject" else "rejected"
    for item in db.query(BugItem).filter_by(bug_id=bid, status="pending").all():
        item.status = child_status
        item.adopt_level = body.adopt_level
        item.final_score = body.final_score
        item.reviewed_by = user.id
        item.reviewed_at = datetime.utcnow()
    if body.adopt_level != "reject":
        # 报告人贡献按终审分加权
        record_contribution(db, b.reporter_id, "bug_accepted", "bug", bid,
                            b.dataset_id, weight=body.final_score)
        # 参与评审且方向与终审一致者 +k
        member_reviews = db.query(CorrectionReview).filter_by(
            target_type="bug", target_id=bid, reviewer_type="member").all()
        for r in member_reviews:
            aligned = (r.acceptability_score >= 5) == (body.final_score >= 5)
            if aligned and r.reviewer_id:
                record_contribution(db, r.reviewer_id, "review_adopted", "bug", bid,
                                    b.dataset_id, weight=3)
    write_audit(db, user.id, "bug.finalize", "bug", bid,
                {"adopt": body.adopt_level, "score": body.final_score,
                 "self_review": self_review})
    db.commit()
    return {"ok": True, "status": b.status}


# ================= bug 证据附件（真实文件上传）=================
from fastapi import UploadFile, File  # noqa: E402
from fastapi.responses import StreamingResponse  # noqa: E402


@router.post("/bugs/{bid}/attachments")
def upload_bug_attachment(bid: int, file: UploadFile = File(...),
                          db: Session = Depends(get_db),
                          user: User = Depends(get_current_user)):
    from ..models.correction import BugAttachment
    from ..services.uploads import save_upload
    from datetime import datetime
    b = db.get(Bug, bid)
    if not b:
        raise HTTPException(404, "勘误不存在")
    if not is_dataset_member(db, b.dataset_id, user):
        raise HTTPException(403, "需为数据集成员")
    meta = save_upload(file, f"bug/{bid}")
    a = BugAttachment(bug_id=bid, uploaded_by=user.id, uploaded_at=datetime.utcnow(), **meta)
    db.add(a); db.commit(); db.refresh(a)
    return {"id": a.id, "file_name": a.file_name, "size": a.size}


@router.get("/bugs/{bid}/attachments")
def list_bug_attachments(bid: int, db: Session = Depends(get_db),
                         user: User = Depends(get_current_user)):
    from ..models.correction import BugAttachment
    rows = db.query(BugAttachment).filter_by(bug_id=bid).all()
    return [{"id": a.id, "file_name": a.file_name, "size": a.size, "mime": a.mime}
            for a in rows]


@router.get("/bug-attachments/{aid}/download")
def download_bug_attachment(aid: int, db: Session = Depends(get_db),
                            user: User = Depends(get_current_user)):
    from ..models.correction import BugAttachment
    from ..core.storage import storage
    a = db.get(BugAttachment, aid)
    if not a:
        raise HTTPException(404, "附件不存在")
    b = db.get(Bug, a.bug_id)
    if not is_dataset_member(db, b.dataset_id, user):
        raise HTTPException(403, "需为数据集成员")
    from ..services.uploads import open_stored_file
    stream = open_stored_file(a.file_path)
    from ..services.downloads import log_download
    from ..models.dataset import Dataset as _DS
    _d = db.get(_DS, b.dataset_id)
    log_download(db, user_id=user.id, source="bug_attachment", dataset_id=b.dataset_id,
                 location_label=(_d.name_zh if _d else "勘误附件"),
                 detail=f"勘误#{b.id} 附件", file_name=a.file_name,
                 link=(f"/#/datasets/{_d.slug}?tab=bugs&bug={b.id}" if _d else ""))
    db.commit()
    from ..services.uploads import attachment_headers
    return StreamingResponse(stream, media_type=a.mime or "application/octet-stream",
                             headers=attachment_headers(a.file_name))


@router.get("/bugs/{bid}")
def bug_detail(bid: int, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    from ..models.correction import BugAttachment
    from ..models.version import DataVersion
    b = db.get(Bug, bid)
    if not b:
        raise HTTPException(404, "勘误不存在")
    reporter = db.get(User, b.reporter_id)
    reviews = db.query(CorrectionReview).filter_by(target_type="bug", target_id=bid).all()
    fin = db.query(CorrectionFinal).filter_by(target_type="bug", target_id=bid).first()
    atts = db.query(BugAttachment).filter_by(bug_id=bid).all()
    fixed_v = db.get(DataVersion, b.fixed_in_version_id) if b.fixed_in_version_id else None
    items = db.query(BugItem).filter_by(bug_id=bid).order_by(BugItem.seq).all()
    item_ai = {}
    for review in db.query(CorrectionReview).filter_by(
            target_type="bug_item", reviewer_type="ai").filter(
            CorrectionReview.target_id.in_([it.id for it in items] or [-1])).all():
        item_ai[review.target_id] = review.comment
    reviewer_count = db.query(CorrectionReview.reviewer_id).filter_by(
        target_type="bug", target_id=bid, reviewer_type="member").filter(
        CorrectionReview.reviewer_id.is_not(None)).distinct().count()
    editor_ids = {it.admin_edited_by for it in items if it.admin_edited_by}
    editors = {
        editor.id: editor.display_name
        for editor in db.query(User).filter(User.id.in_(editor_ids or {-1})).all()
    }
    recommended_uid_var = _uid_var(db, b.dataset_id)
    correction_type = "unstructured" if b.bug_type == "unstructured" else "structured"
    return {"id": b.id, "dataset_id": b.dataset_id, "officer_id": b.officer_id,
            "term_id": b.term_id, "current_value": b.current_value,
            "suggested_value": b.suggested_value, "description_zh": b.description_zh,
            "correction_type": correction_type,
            "manual_only": correction_type == "unstructured",
            "evidence": b.evidence, "status": _effective_bug_status(items, b.status),
            "created_at": china_iso(b.created_at), "reviewer_count": reviewer_count,
            "reporter": {"id": b.reporter_id, "name": reporter.display_name if reporter else ""},
            "can_delete": (
                b.reporter_id == user.id and b.status == "pending"
                and all(it.status == "pending" for it in items)
            ),
            "fixed_in_version": fixed_v.version_id if fixed_v else None,
            "reviews": [{"reviewer_type": r.reviewer_type, "reviewer_id": r.reviewer_id,
                         "score": r.acceptability_score, "comment": r.comment} for r in reviews],
            "final": ({"adopt_level": fin.adopt_level, "final_score": fin.final_score,
                       "comment": fin.comment} if fin else None),
            "items": [{"id": it.id, "seq": it.seq,
                       "uid_var": it.uid_var or recommended_uid_var,
                       "uses_alternative_id": bool(
                           it.uid_var and it.uid_var != recommended_uid_var),
                       "manual_only": bool(it.is_new_officer),
                       "uid_value": it.uid_value,
                       "var_name": it.var_name, "current_value": it.current_value,
                       "suggested_value": it.suggested_value, "reason": it.reason,
                       "evidence": it.evidence,
                       "is_new_officer": bool(it.is_new_officer),
                       "status": it.status, "ai_score": it.ai_score,
                       "ai_reason": item_ai.get(it.id),
                       "final_score": it.final_score, "adopt_level": it.adopt_level,
                       "admin_modified": it.original_uid_value is not None,
                       "original": ({
                           "uid_value": it.original_uid_value,
                           "var_name": it.original_var_name,
                           "current_value": it.original_current_value,
                           "suggested_value": it.original_suggested_value,
                           "reason": it.original_reason,
                       } if it.original_uid_value is not None else None),
                       "admin_editor": ({
                           "id": it.admin_edited_by,
                           "name": editors.get(it.admin_edited_by, "管理员"),
                       } if it.admin_edited_by else None),
                       "admin_edited_at": china_iso(it.admin_edited_at)}
                      for it in items],
            "attachments": [{"id": a.id, "file_name": a.file_name, "size": a.size} for a in atts]}


# ============ 逐条子项：AI 评分 / 管理员终审 ============
@router.post("/bug-items/{iid}/ai-review")
def ai_review_item(iid: int, db: Session = Depends(get_db),
                   user: User = Depends(get_current_user)):
    it = db.get(BugItem, iid)
    if not it:
        raise HTTPException(404, "勘误项不存在")
    if not is_dataset_member(db, it.dataset_id, user):
        raise HTTPException(403, "需为数据集成员")
    if not (it.reason or "").strip() or not (it.evidence or "").strip():
        raise HTTPException(400, "该勘误项缺少独立的说明或证据，补充完整后才能进行 AI 评分")
    prompt = (f"变量：{it.var_name}\n唯一 ID：{it.uid_value}\n"
              f"当前值：{it.current_value or ''}\n建议值：{it.suggested_value or ''}\n"
              f"修改说明：{it.reason}\n证据：{it.evidence}")
    score, reason = _run_ai_review(prompt)
    it.ai_score = score
    review = db.query(CorrectionReview).filter_by(
        target_type="bug_item", target_id=iid, reviewer_type="ai").first()
    if not review:
        review = CorrectionReview(target_type="bug_item", target_id=iid,
                                  reviewer_type="ai")
        db.add(review)
    review.acceptability_score = score
    review.comment = reason
    db.commit()
    return {"ai_score": it.ai_score, "reason": reason}


@router.post("/bug-items/{iid}/finalize")
def finalize_item(iid: int, body: FinalizeIn, db: Session = Depends(get_db),
                  user: User = Depends(get_current_user)):
    it = db.get(BugItem, iid)
    if not it:
        raise HTTPException(404, "勘误项不存在")
    if not is_dataset_admin(db, it.dataset_id, user):
        raise HTTPException(403, "仅数据集管理员可审核")
    if it.status != "pending":
        raise HTTPException(400, "该勘误项已处理")
    if body.adopt_level == "partial":
        raise HTTPException(
            409, "部分采纳不能直接确认；请先进入修改页面并实际修改勘误内容")
    b = db.get(Bug, it.bug_id)
    # 自审校验：本人提交且存在其他管理员时须由他人审
    if b and b.reporter_id == user.id and count_dataset_admins(db, it.dataset_id) > 1:
        raise HTTPException(403, "这是你本人提交的勘误，请由另一名管理员审核")
    it.adopt_level = body.adopt_level
    it.final_score = body.final_score
    it.status = "accepted" if body.adopt_level != "reject" else "rejected"
    it.reviewed_by = user.id; it.reviewed_at = datetime.utcnow()
    if body.adopt_level != "reject" and b:
        record_contribution(db, b.reporter_id, "bug_accepted", "bug_item", iid,
                            it.dataset_id, weight=body.final_score)
    if b:
        _sync_bug_status(db, b)
        b.reviewed_by = user.id
        b.reviewed_at = datetime.utcnow()
    write_audit(db, user.id, "bug.item.finalize", "bug_item", iid,
                {"adopt": body.adopt_level, "score": body.final_score})
    db.commit()
    return {"ok": True, "status": it.status}


@router.post("/bug-items/{iid}/finalize-partial")
def finalize_item_partial(iid: int, body: PartialFinalizeIn,
                          db: Session = Depends(get_db),
                          user: User = Depends(get_current_user)):
    """管理员部分采纳：必须先修改，保留原投稿后才进入已采纳未修改。"""
    it = db.get(BugItem, iid)
    if not it:
        raise HTTPException(404, "勘误项不存在")
    if not is_dataset_admin(db, it.dataset_id, user):
        raise HTTPException(403, "仅数据集管理员可审核")
    if it.status != "pending":
        raise HTTPException(400, "该勘误项已处理")
    bug = db.get(Bug, it.bug_id)
    if bug and bug.reporter_id == user.id and count_dataset_admins(db, it.dataset_id) > 1:
        raise HTTPException(403, "这是你本人提交的勘误，请由另一名管理员审核")

    context = _uid_context(db, it.dataset_id, body.uid_var or it.uid_var)
    uid_check = _check_uid(context, body.uid_value)
    if not uid_check["exists"] and not body.confirm_new_officer:
        raise HTTPException(
            409, f"最新原始版本中没有 {context['unique_id_var']}={uid_check['value']}。"
                 "若确为新增主体或新增记录，请明确勾选后再确认")
    variable = db.query(Variable).filter_by(
        dataset_id=it.dataset_id, var_name=body.var_name, enabled=True).first()
    if not variable:
        raise HTTPException(400, "请选择本数据集当前有效的勘误变量")
    if variable.var_name in {
            context["unique_id_var"], context["recommended_uid_var"]}:
        raise HTTPException(400, "管理员推荐或本次选用的 ID 变量本身不能作为勘误修改对象")

    old_values = (
        _match_text(it.uid_value), _match_text(it.var_name),
        _match_text(it.current_value), _match_text(it.suggested_value),
        _match_text(it.reason),
    )
    new_values = (
        _match_text(uid_check["value"]), _match_text(variable.var_name),
        _match_text(body.current_value), _match_text(body.suggested_value),
        _match_text(body.reason),
    )
    if new_values == old_values:
        raise HTTPException(400, "选择“部分采纳”后必须至少实际修改一项内容")

    duplicate = _duplicate_index(
        db, it.dataset_id, context, exclude_bug_id=it.bug_id,
        exclude_item_id=it.id,
    ).get(_bug_signature(
        context, uid_check["value"], variable.var_name,
        body.current_value, body.suggested_value,
    ))
    if duplicate:
        raise HTTPException(409, "修改后仍是重复勘误：" + _duplicate_message(duplicate))

    it.original_uid_value = it.uid_value
    it.original_var_name = it.var_name
    it.original_current_value = it.current_value
    it.original_suggested_value = it.suggested_value
    it.original_reason = it.reason
    it.uid_var = context["unique_id_var"]
    it.uid_value = uid_check["value"]
    it.var_name = variable.var_name
    it.current_value = body.current_value
    it.suggested_value = body.suggested_value
    it.reason = body.reason
    it.is_new_officer = not uid_check["exists"]
    it.status = "accepted"
    it.adopt_level = "partial"
    it.final_score = body.final_score
    it.reviewed_by = user.id
    it.reviewed_at = datetime.utcnow()
    it.admin_edited_by = user.id
    it.admin_edited_at = datetime.utcnow()

    sibling_count = db.query(BugItem).filter_by(bug_id=it.bug_id).count()
    if bug:
        # 新流程一条 Bug 只有一个子项；同步父记录供列表、AI 与历史接口继续使用。
        if sibling_count == 1:
            bug.officer_id = uid_check["value"]
            bug.variable_id = variable.id
            bug.current_value = body.current_value
            bug.suggested_value = body.suggested_value
            bug.description_zh = body.reason
            final = db.query(CorrectionFinal).filter_by(
                target_type="bug", target_id=bug.id).first()
            if not final:
                final = CorrectionFinal(target_type="bug", target_id=bug.id)
                db.add(final)
            final.decided_by = user.id
            final.adopt_level = "partial"
            final.final_score = body.final_score
            final.comment = body.comment or "管理员部分采纳并修改勘误内容"
            final.decided_at = datetime.utcnow()
        _sync_bug_status(db, bug)
        bug.reviewed_by = user.id
        bug.reviewed_at = datetime.utcnow()
        record_contribution(
            db, bug.reporter_id, "bug_accepted", "bug_item", iid,
            it.dataset_id, weight=body.final_score,
        )
    write_audit(db, user.id, "bug.item.finalize.partial", "bug_item", iid, {
        "score": body.final_score,
        "original": {
            "uid_value": old_values[0], "var_name": old_values[1],
            "current_value": old_values[2], "suggested_value": old_values[3],
            "reason": old_values[4],
        },
        "modified": {
            "uid_var": context["unique_id_var"],
            "uid_value": new_values[0], "var_name": new_values[1],
            "current_value": new_values[2], "suggested_value": new_values[3],
            "reason": new_values[4],
        },
    })
    db.commit()
    return {"ok": True, "status": it.status, "adopt_level": "partial"}


# ============ 一键把已采纳勘误应用到上一版数据，生成新版本 ============
def _accepted_release_material(db: Session, dataset_id: int, unique_id_var: str,
                               base_version_id: int):
    """生成勘误发版所需的结构化修改项、文字说明和实际修改代码。"""
    from ..services.data_ops import apply_corrections_script
    items = db.query(BugItem).filter_by(dataset_id=dataset_id, status="accepted").filter(
        BugItem.applied_in_version.is_(None)).order_by(BugItem.bug_id, BugItem.seq).all()
    payload = [{"seq": it.id, "bug_id": it.bug_id, "item_seq": it.seq,
                "uid_var": it.uid_var or unique_id_var,
                "uid_value": it.uid_value, "var_name": it.var_name,
                "current_value": it.current_value, "suggested_value": it.suggested_value,
                "reason": it.reason, "evidence": it.evidence,
                "is_new_officer": bool(it.is_new_officer)} for it in items]
    auto_payload = [it for it in payload if not it["is_new_officer"]]
    manual_payload = [it for it in payload if it["is_new_officer"]]
    script = apply_corrections_script(payload, unique_id_var) if payload else ""
    lines = [f"本版本自动应用 {len(auto_payload)} 条已采纳勘误；"
             f"{len(manual_payload)} 条勘误保留为人工处理："]
    for it in payload:
        if it["is_new_officer"]:
            mode = "人工处理·新增官员/记录"
        elif it["uid_var"] != unique_id_var:
            mode = f"自动应用·非推荐 ID（{it['uid_var']}）"
        else:
            mode = "自动应用"
        reason = f"；说明：{it['reason']}" if it.get("reason") else ""
        lines.append(
            f"- [{mode}] 勘误 #{it['bug_id']}·第 {it['item_seq']} 项："
            f"{it['uid_var']}={it['uid_value']}，"
            f"{it['var_name']} 由「{it.get('current_value') or ''}」改为「{it.get('suggested_value') or ''}」{reason}")
    fingerprint = hashlib.sha256(json.dumps(
        {"dataset_id": dataset_id, "base_version_id": base_version_id,
         "unique_id_var": unique_id_var, "items": payload},
        ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")).hexdigest()
    return items, auto_payload, manual_payload, script, "\n".join(lines), fingerprint


@router.get("/datasets/{slug}/corrections-release-preview")
def corrections_release_preview(slug: str, base_version_id: int | None = None,
                                db: Session = Depends(get_db),
                                user: User = Depends(get_current_user)):
    """一键发版弹窗默认说明：同时给出人可读摘要与可复核修改代码。"""
    d = _ds(db, slug)
    if not is_dataset_admin(db, d.id, user):
        raise HTTPException(403, "仅数据集管理员可查看已采纳勘误的发版说明")
    cfg = db.get(DatasetDataConfig, d.id)
    uidv = cfg.unique_id_var if cfg else None
    if not uidv:
        raise HTTPException(400, "请先在「数据处理设置」指定唯一ID变量，再生成勘误发版说明")
    base = db.get(DataVersion, base_version_id) if base_version_id else _latest_raw_version(
        db, d.id)
    extra = db.get(VersionExtra, base.id) if base else None
    if (not base or base.dataset_id != d.id or not base.data_file_path or
            (extra.data_kind if extra else "raw") != "raw"):
        raise HTTPException(400, "请选择本数据集带数据文件的原始版本作为基准版本")
    _, auto, manual, script, changelog, preview_hash = _accepted_release_material(
        db, d.id, uidv, base.id)
    if not auto and not manual:
        raise HTTPException(400, "没有待应用的已采纳勘误项；请先完成勘误终审")
    return {"count": len(auto) + len(manual), "auto_count": len(auto),
            "manual_count": len(manual), "manual_items": manual,
            "base_version_id": base.id, "changelog_zh": changelog,
            "script": script, "preview_hash": preview_hash}


@router.post("/datasets/{slug}/apply-corrections")
def apply_corrections_endpoint(slug: str, base_version_id: int = Form(...),
                               new_version_id: str = Form(...), changelog_zh: str = Form(""),
                               preview_hash: str = Form(...),
                               db: Session = Depends(get_db),
                               user: User = Depends(get_current_user)):
    from ..services.data_ops import apply_corrections, CorrectionPreconditionError
    d = _ds(db, slug)
    if not is_dataset_admin(db, d.id, user):
        raise HTTPException(403, "仅数据集管理员可应用勘误发版")
    if get_settings(db, d.id).is_closed:
        raise HTTPException(400, "数据集已关闭")
    base = db.get(DataVersion, base_version_id)
    if not base or base.dataset_id != d.id or not base.data_file_path:
        raise HTTPException(404, "基准版本不存在或无数据文件")
    base_extra = db.get(VersionExtra, base.id)
    if (base_extra.data_kind if base_extra else "raw") != "raw":
        raise HTTPException(400, "勘误只能应用到原始数据版本")
    if not (new_version_id or "").strip():
        raise HTTPException(400, "请填写新版本号")
    if db.query(DataVersion).filter_by(dataset_id=d.id, version_id=new_version_id).first():
        raise HTTPException(400, f"版本 {new_version_id} 已存在")
    cfg = db.get(DatasetDataConfig, d.id)
    uidv = cfg.unique_id_var if cfg else None
    if not uidv:
        raise HTTPException(400, "请先在数据设置里指定唯一ID变量，才能按ID定位修改")
    items, payload, manual, script, default_changelog, current_hash = (
        _accepted_release_material(db, d.id, uidv, base.id))
    if current_hash != preview_hash:
        raise HTTPException(409, "已采纳勘误或基准版本在预览后发生变化，请重新打开并审阅代码")
    if not payload and manual:
        return {"generated": "script", "script": script, "manual_remaining": len(manual),
                "note": "待处理项均为新增主体或新增记录，系统不会自动追加不完整记录。"
                        "请人工处理后再发布原始版本。"}
    if not payload:
        raise HTTPException(400, "没有待应用的已采纳勘误项")
    try:
        new_bytes, source, generated_script, applied_ids = apply_corrections(
            base.data_file_path, payload, uidv, cfg.script_only if cfg else False)
    except CorrectionPreconditionError as exc:
        raise HTTPException(409, f"数据与勘误预期不一致，已中止整批应用：{exc}") from exc
    if new_bytes is None:
        write_audit(db, user.id, "dataset.apply_corrections.script", "dataset", d.id)
        return {"generated": "script", "script": script,
                "manual_remaining": len(manual),
                "note": "数据过大或设为仅脚本模式：请在本地运行脚本改好数据后，用「发布新版本·原始」上传。"}
    key = f"versions/{d.slug}/{new_version_id}/data.dta"
    from ..services.uploads import save_stored_file
    save_stored_file(key, io.BytesIO(new_bytes))
    db.query(DataVersion).filter_by(dataset_id=d.id, is_current=True).update(
        {"is_current": False, "valid_to": datetime.utcnow()})
    audit_changelog = ((changelog_zh or "").strip() or default_changelog)
    audit_changelog += "\n\n管理员审阅的实际修改代码：\n" + script.rstrip()
    v = DataVersion(dataset_id=d.id, version_id=new_version_id, based_on_version=base.version_id,
                    release_date=datetime.utcnow(), data_file_path=key,
                    changelog_zh=audit_changelog,
                    created_by=user.id, is_current=True, valid_from=datetime.utcnow())
    db.add(v); db.flush()
    db.add(VersionExtra(version_id=v.id, data_kind="raw", generated=source))
    d.current_version_id = v.id
    applied_set = set(applied_ids)
    for it in items:
        if it.id in applied_set:
            it.status = "fixed"; it.applied_in_version = v.id
    affected_bug_ids = {it.bug_id for it in items if it.id in applied_set}
    for bug_id in affected_bug_ids:
        bug = db.get(Bug, bug_id)
        if bug:
            _sync_bug_status(db, bug, v.id)
    write_audit(db, user.id, "dataset.apply_corrections.server", "dataset", d.id,
                {"version": new_version_id, "applied": len(applied_ids)})
    db.commit()
    return {"generated": "server", "id": v.id, "version_id": new_version_id,
            "applied": len(applied_ids), "manual_remaining": len(manual)}
