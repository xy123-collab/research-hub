import io
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from ..core.db import get_db
from ..core.storage import storage
from ..core.ai_client import ai_client
from ..core.permissions import (get_current_user, is_super_admin, is_dataset_member,
                                is_dataset_admin, dataset_membership, has_dataset_perm,
                                count_dataset_admins, get_settings, active_grants,
                                DS_ADMIN_ROLES, DS_LEAD_ROLES, dataset_lead_id,
                                is_dataset_lead)
from ..core.audit import write_audit, record_contribution
from ..models.user import User
from ..models.dataset import (Dataset, DatasetMember, JoinRequest, Variable,
                              DatasetGroupRequest)
from ..models.access import (DatasetGrant, DatasetSettings, DownloadRequest,
                             VersionCandidate, CodebookDraft, GRANTABLE_PERMS,
                             PERM_LABELS_ZH, DOWNLOAD_POLICIES)
from ..models.curation import (VersionExtra, DatasetDataConfig, VariableMaskRule,
                               DATA_KINDS, MASK_ACTIONS)
from ..models.version import DataVersion, DownloadLog
from ..models.literature import (LitTopic, LitRef, Publication, DatasetSummary)
from ..models.group import ResearchGroup, Charter
from ..models.correction import Bug
from ..models.code import CodeScript
from ..models.governance import VerifyFlag
from ..schemas.models import (VersionIn, LitRefIn, DatasetIn, LitBatchIn, CitationTextIn,
                              AiSummaryIn, AiHintIn)

router = APIRouter(tags=["datasets"])


def _get_ds(db, slug) -> Dataset:
    d = db.query(Dataset).filter_by(slug=slug, is_deleted=False).first()
    if not d:
        raise HTTPException(404, "数据集不存在")
    return d


def _ds_card(db, d, user):
    """首页/发现页用的数据集卡片：带课题组、当前版本、协作活跃度信号（只读）。"""
    n = db.query(DatasetMember).filter_by(dataset_id=d.id).count()
    g = db.get(ResearchGroup, d.group_id) if d.group_id else None
    cur = db.query(DataVersion).filter_by(dataset_id=d.id, is_current=True).first()
    pending = db.query(Bug).filter_by(dataset_id=d.id, status="pending").count()
    open_flags = db.query(VerifyFlag).filter_by(dataset_id=d.id, status="open").count()
    m = dataset_membership(db, d.id, user.id)
    return {"id": d.id, "slug": d.slug, "name_zh": d.name_zh, "name_en": d.name_en,
            "icon": d.icon, "desc_zh": d.desc_zh, "group_id": d.group_id,
            "group_name": g.name_zh if g else "", "group_slug": g.slug if g else "",
            "member_count": n, "is_sensitive": d.is_sensitive,
            "current_version": cur.version_id if cur else None,
            "pending_bugs": pending, "open_flags": open_flags,
            "is_member": m is not None, "my_role": m.ds_role if m else None}


@router.get("/datasets")
def wall(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """全平台公开数据集墙（发现用）。字段向后兼容，仅新增协作信号。"""
    return [_ds_card(db, d, user) for d in
            db.query(Dataset).filter_by(is_deleted=False, is_public=True).all()]


def _recent_events(db, d):
    """数据集近期消息：最新版本 + 最新勘误，供首页直观展示。"""
    ev = []
    v = db.query(DataVersion).filter_by(dataset_id=d.id).order_by(DataVersion.id.desc()).first()
    if v:
        ev.append({"type": "version", "text": f"发布 {v.version_id}",
                   "at": str(v.release_date)[:10] if v.release_date else ""})
    b = db.query(Bug).filter_by(dataset_id=d.id).order_by(Bug.id.desc()).first()
    if b:
        ev.append({"type": "bug", "text": f"勘误 #{b.id}（{b.status}）",
                   "at": str(b.reviewed_at)[:10] if b.reviewed_at else ""})
    return ev


@router.get("/datasets/mine")
def my_datasets(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """我参与（成员/维护者/发起人）的数据集，按待办协作量排序，供首页直达协作。"""
    ds_ids = [m.dataset_id for m in
              db.query(DatasetMember).filter_by(user_id=user.id).all()]
    cards = []
    for d in db.query(Dataset).filter(Dataset.id.in_(ds_ids or [-1]),
                                       Dataset.is_deleted == False).all():
        card = _ds_card(db, d, user)
        card["recent"] = _recent_events(db, d)
        cards.append(card)
    cards.sort(key=lambda c: (c["pending_bugs"] + c["open_flags"]), reverse=True)
    return cards


@router.post("/datasets")
def create_standalone_dataset(body: DatasetIn, db: Session = Depends(get_db),
                              user: User = Depends(get_current_user)):
    """独立创建数据集（不归属任何课题组）。创建者即发起人/数据集管理员。"""
    if db.query(Dataset).filter_by(slug=body.slug).first():
        raise HTTPException(400, "数据集 slug 已存在")
    d = Dataset(group_id=None, founder_id=user.id, **body.model_dump())
    db.add(d); db.flush()
    db.add(DatasetMember(dataset_id=d.id, user_id=user.id, ds_role="founder",
                         joined_at=datetime.utcnow(), approved_by=user.id))
    db.add(Charter(scope="dataset", ref_id=d.id,
                   body_zh="（请数据集发起人编辑本数据集公约）", version=1, updated_by=user.id))
    record_contribution(db, user.id, "dataset_founder", "dataset", d.id, d.id, weight=30)
    write_audit(db, user.id, "dataset.create_standalone", "dataset", d.id)
    db.commit()
    return {"id": d.id, "slug": d.slug}


@router.get("/datasets/{slug}/activity")
def dataset_activity(slug: str, kind: str = "all", db: Session = Depends(get_db),
                     user: User = Depends(get_current_user)):
    """更新记录时间线：版本发布 / 勘误 / 处理代码，可用 kind 过滤（all|version|bug|code）。"""
    d = _get_ds(db, slug)
    items = []
    if kind in ("all", "version"):
        for v in db.query(DataVersion).filter_by(dataset_id=d.id).all():
            items.append({"type": "version", "title": f"发布版本 {v.version_id}",
                          "detail": v.changelog_zh or "", "ref": v.version_id,
                          "at": str(v.release_date) if v.release_date else None, "sort": v.id})
    if kind in ("all", "bug"):
        for b in db.query(Bug).filter_by(dataset_id=d.id).all():
            items.append({"type": "bug", "title": f"勘误 #{b.id}（{b.status}）",
                          "detail": b.description_zh or "", "ref": b.id,
                          "at": str(b.reviewed_at) if b.reviewed_at else None, "sort": b.id})
    if kind in ("all", "code"):
        for c in db.query(CodeScript).filter_by(dataset_id=d.id).all():
            items.append({"type": "code", "title": f"处理代码 {c.filename}",
                          "detail": c.title_zh or "", "ref": c.id, "at": None, "sort": c.id})
    items.sort(key=lambda x: (x["at"] is not None, x["at"] or "", x["sort"]), reverse=True)
    return items


@router.post("/datasets/{slug}/attach-request")
def attach_request(slug: str, group_slug: str, db: Session = Depends(get_db),
                   user: User = Depends(get_current_user)):
    """数据集管理员申请把（独立）数据集并入某课题组，待该组管理员审批。"""
    d = _get_ds(db, slug)
    if not is_dataset_admin(db, d.id, user):
        raise HTTPException(403, "仅数据集发起人/管理员可申请归属")
    if d.group_id:
        raise HTTPException(400, "数据集已归属某课题组，请先申请移出")
    g = db.query(ResearchGroup).filter_by(slug=group_slug, is_deleted=False).first()
    if not g:
        raise HTTPException(404, "课题组不存在")
    if db.query(DatasetGroupRequest).filter_by(dataset_id=d.id, status="pending").first():
        raise HTTPException(400, "已有待处理的归属申请")
    r = DatasetGroupRequest(dataset_id=d.id, group_id=g.id, kind="attach",
                            requested_by=user.id, status="pending", created_at=datetime.utcnow())
    db.add(r); db.commit()
    return {"id": r.id, "status": "pending"}


@router.post("/datasets/{slug}/detach-request")
def detach_request(slug: str, db: Session = Depends(get_db),
                   user: User = Depends(get_current_user)):
    """数据集管理员申请把数据集移出当前课题组（变为独立），需该组管理员同意。"""
    d = _get_ds(db, slug)
    if not is_dataset_admin(db, d.id, user):
        raise HTTPException(403, "仅数据集发起人/管理员可申请移出")
    if not d.group_id:
        raise HTTPException(400, "数据集当前不属于任何课题组")
    if db.query(DatasetGroupRequest).filter_by(dataset_id=d.id, status="pending").first():
        raise HTTPException(400, "已有待处理的归属申请")
    r = DatasetGroupRequest(dataset_id=d.id, group_id=d.group_id, kind="detach",
                            requested_by=user.id, status="pending", created_at=datetime.utcnow())
    db.add(r); db.commit()
    return {"id": r.id, "status": "pending"}


@router.get("/datasets/{slug}")
def detail(slug: str, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    d = _get_ds(db, slug)
    member = is_dataset_member(db, d.id, user)
    founder = db.get(User, d.founder_id)
    cur = db.query(DataVersion).filter_by(dataset_id=d.id, is_current=True).first()
    lead_id = dataset_lead_id(db, d.id)
    members = db.query(DatasetMember).filter_by(dataset_id=d.id).all()
    approvals = []
    for m in members:
        u = db.get(User, m.user_id)
        approvals.append({"user_id": m.user_id, "name": u.display_name if u else "",
                          "ds_role": m.ds_role, "is_lead": m.user_id == lead_id,
                          "is_admin": m.ds_role in DS_ADMIN_ROLES,
                          "joined_at": str(m.joined_at), "approved_by": m.approved_by})
    charter = db.query(Charter).filter_by(scope="dataset", ref_id=d.id).order_by(
        Charter.version.desc()).first()
    pubs = db.query(Publication).filter_by(dataset_id=d.id).all()
    grp = db.get(ResearchGroup, d.group_id) if d.group_id else None
    pend = db.query(DatasetGroupRequest).filter_by(
        dataset_id=d.id, status="pending").first()
    pend_group = db.get(ResearchGroup, pend.group_id) if pend else None
    is_admin = is_dataset_admin(db, d.id, user)
    st = get_settings(db, d.id)
    # 我拥有的单独授权码（成员用于前端按钮显隐；管理员视为全部）
    if is_admin:
        my_perms = list(GRANTABLE_PERMS)
    elif member:
        my_perms = sorted({g.perm for g in active_grants(db, d.id, user.id)} |
                          set((dataset_membership(db, d.id, user.id).granted_perms_json) or []))
    else:
        my_perms = []
    # 我是否已有有效的下载申请（审批后下载策略下用）
    my_dl_req = db.query(DownloadRequest).filter_by(
        dataset_id=d.id, user_id=user.id).order_by(DownloadRequest.id.desc()).first()
    return {
        "id": d.id, "slug": d.slug, "name_zh": d.name_zh, "name_en": d.name_en,
        "desc_zh": d.desc_zh, "icon": d.icon, "is_sensitive": d.is_sensitive,
        "group": ({"slug": grp.slug, "name_zh": grp.name_zh} if grp else None),
        "pending_group_request": ({"kind": pend.kind, "group_name": pend_group.name_zh
                                   if pend_group else "", "status": pend.status}
                                  if pend else None),
        "founder": {"id": d.founder_id, "name": founder.display_name if founder else "",
                    "contact": d.founder_contact},
        "is_member": member, "is_admin": is_admin,
        "is_lead": lead_id == user.id, "lead_id": lead_id,
        "unique_id_var": (_data_config(db, d.id).unique_id_var),
        "charter": ({"id": charter.id, "body_zh": charter.body_zh,
                     "version": charter.version} if charter else None),
        "my_perms": my_perms,
        "settings": {"download_policy": st.download_policy,
                     "history_visible": st.history_visible,
                     "history_downloadable": st.history_downloadable,
                     "analysis_open": st.analysis_open, "is_closed": st.is_closed},
        "my_download_request": ({"status": my_dl_req.status} if my_dl_req else None),
        "current_version": ({"id": cur.id, "version_id": cur.version_id,
                             "changelog_zh": cur.changelog_zh} if cur else None),
        "members": approvals,
        "publications": [{"title": p.title, "venue": p.venue, "year": p.year,
                          "url": p.url} for p in pubs],
    }


# -------- variables / codebook（非成员亦可看）--------
@router.get("/datasets/{slug}/variables")
def variables(slug: str, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    d = _get_ds(db, slug)
    vs = db.query(Variable).filter_by(dataset_id=d.id, enabled=True).all()
    return [{"id": v.id, "var_name": v.var_name, "group_code": v.group_code,
             "label_zh": v.label_zh, "label_en": v.label_en} for v in vs]


# -------- versions --------
@router.get("/datasets/{slug}/versions")
def list_versions(slug: str, db: Session = Depends(get_db),
                  user: User = Depends(get_current_user)):
    d = _get_ds(db, slug)
    vs = db.query(DataVersion).filter_by(dataset_id=d.id).order_by(
        DataVersion.id.desc()).all()
    out = []
    for v in vs:
        ex = db.get(VersionExtra, v.id)
        out.append({"id": v.id, "version_id": v.version_id, "based_on": v.based_on_version,
                    "release_date": str(v.release_date), "changelog_zh": v.changelog_zh,
                    "is_current": v.is_current, "has_data": bool(v.data_file_path),
                    "data_kind": ex.data_kind if ex else "raw",
                    "masked_source": ex.masked_source_version if ex else None})
    return out


@router.post("/datasets/{slug}/versions")
async def publish_version(slug: str, version_id: str = Form(...),
                          based_on_version: str = Form(""),
                          changelog_zh: str = Form(""), changelog_en: str = Form(""),
                          fixed_bug_ids: str = Form(""),  # 逗号分隔：本次修复的已采纳勘误
                          data_kind: str = Form("raw"),    # raw|masked|sample
                          data_file: UploadFile | None = File(None),
                          codebook_file: UploadFile | None = File(None),
                          db: Session = Depends(get_db),
                          user: User = Depends(get_current_user)):
    d = _get_ds(db, slug)
    if not is_dataset_admin(db, d.id, user):
        raise HTTPException(403, "仅发起人/管理员可发版")
    if get_settings(db, d.id).is_closed:
        raise HTTPException(400, "数据集已关闭，不再发布新版本")
    if data_kind not in DATA_KINDS:
        raise HTTPException(400, "数据分类须为 原始/脱敏/样例")
    # 版本不可覆盖：同 (dataset, version_id) 不允许重复
    if db.query(DataVersion).filter_by(dataset_id=d.id, version_id=version_id).first():
        raise HTTPException(400, f"版本 {version_id} 已存在，版本不可覆盖")
    data_key = codebook_key = None
    if data_file:
        if not data_file.filename.endswith(".dta"):
            raise HTTPException(400, "数据文件仅支持 .dta")
        data_key = f"versions/{d.slug}/{version_id}/{data_file.filename}"
        storage.save(data_key, data_file.file)
    if codebook_file:
        codebook_key = f"versions/{d.slug}/{version_id}/{codebook_file.filename}"
        storage.save(codebook_key, codebook_file.file)
    # 样例数据不参与"当前推荐版"迭代（公开、独立、不迭代）；原始/脱敏才竞争 current
    is_iterating = data_kind in ("raw", "masked")
    if is_iterating:
        db.query(DataVersion).filter_by(dataset_id=d.id, is_current=True).update(
            {"is_current": False, "valid_to": datetime.utcnow()})
    v = DataVersion(dataset_id=d.id, version_id=version_id, based_on_version=based_on_version,
                    release_date=datetime.utcnow(), data_file_path=data_key,
                    codebook_file_path=codebook_key, changelog_zh=changelog_zh,
                    changelog_en=changelog_en, created_by=user.id, is_current=is_iterating,
                    valid_from=datetime.utcnow())
    db.add(v); db.flush()
    db.add(VersionExtra(version_id=v.id, data_kind=data_kind))
    if is_iterating:
        d.current_version_id = v.id
    # 核心闭环最后一环：本次修复的已采纳勘误标 fixed + fixed_in_version_id
    from ..models.correction import Bug
    fixed = []
    for raw in [x.strip() for x in fixed_bug_ids.split(",") if x.strip()]:
        bug = db.get(Bug, int(raw))
        if bug and bug.dataset_id == d.id and bug.status == "accepted":
            bug.status = "fixed"; bug.fixed_in_version_id = v.id
            fixed.append(bug.id)
    write_audit(db, user.id, "version.publish", "dataset", d.id,
                {"version": version_id, "fixed_bugs": fixed})
    db.commit()
    return {"id": v.id, "version_id": version_id, "fixed_bugs": fixed}


@router.patch("/datasets/{slug}")
def edit_dataset(slug: str, body: dict, db: Session = Depends(get_db),
                 user: User = Depends(get_current_user)):
    """编辑数据集元信息（名称/简介/图标/联系方式/敏感标记）。"""
    d = _get_ds(db, slug)
    if not is_dataset_admin(db, d.id, user):
        raise HTTPException(403, "仅发起人/管理员可编辑")
    allowed = {"name_zh", "name_en", "desc_zh", "desc_en", "icon",
               "founder_contact", "is_sensitive"}
    for k, val in body.items():
        if k in allowed:
            setattr(d, k, val)
    write_audit(db, user.id, "dataset.edit", "dataset", d.id)
    db.commit()
    return {"ok": True}


@router.patch("/datasets/{slug}/versions/{vid}")
def edit_version(slug: str, vid: int, body: dict, db: Session = Depends(get_db),
                 user: User = Depends(get_current_user)):
    """仅可改元数据 changelog（版本文件不可覆盖）。"""
    d = _get_ds(db, slug)
    if not is_dataset_admin(db, d.id, user):
        raise HTTPException(403, "仅发起人/管理员可编辑")
    v = db.get(DataVersion, vid)
    if not v or v.dataset_id != d.id:
        raise HTTPException(404, "版本不存在")
    for k in ("changelog_zh", "changelog_en"):
        if k in body:
            setattr(v, k, body[k])
    db.commit()
    return {"ok": True}


@router.delete("/datasets/{slug}/members/{uid}")
def remove_member(slug: str, uid: int, db: Session = Depends(get_db),
                  user: User = Depends(get_current_user)):
    d = _get_ds(db, slug)
    if not is_dataset_admin(db, d.id, user):
        raise HTTPException(403, "仅发起人/管理员可移除成员")
    m = dataset_membership(db, d.id, uid)
    if not m:
        raise HTTPException(404, "该用户不是成员")
    if m.ds_role in DS_LEAD_ROLES:
        raise HTTPException(400, "不能移除总管理员；请先转让总管理员身份")
    db.delete(m)
    # 连带撤销其所有单独授权
    db.query(DatasetGrant).filter_by(dataset_id=d.id, user_id=uid).update({"revoked": True})
    write_audit(db, user.id, "dataset.member.remove", "dataset", d.id, {"user": uid})
    db.commit()
    return {"ok": True}


def _has_approved_download(db, dataset_id, user_id, version_id) -> bool:
    """审批后下载：存在已批准且未过期、覆盖该版本的下载申请。"""
    now = datetime.utcnow()
    for r in db.query(DownloadRequest).filter_by(
            dataset_id=dataset_id, user_id=user_id, status="approved").all():
        if r.valid_to and now > r.valid_to:
            continue
        if r.scope_version and version_id and r.scope_version != version_id:
            continue
        return True
    return False


def check_download(db, d, v, user):
    """返回 (允许?, 权限来源/原因)。落实 原则三 与 七节 下载权限类型。"""
    # 样例数据：公开，所有登录用户可下载
    ex = db.get(VersionExtra, v.id)
    if ex and ex.data_kind == "sample":
        return True, "sample"
    admin = is_dataset_admin(db, d.id, user)
    if admin:
        return True, "admin"
    if not is_dataset_member(db, d.id, user):
        return False, "非成员不能下载原始数据，请先申请加入该数据集"
    st = get_settings(db, d.id)
    ver = v.version_id
    # 历史版本单独控制（六节 查看/下载历史版本 授权）
    if not v.is_current:
        if has_dataset_perm(db, d.id, user, "download.history", version=ver):
            return True, "grant:download.history"
        if st.history_downloadable:
            return True, "policy:history_downloadable"
        return False, "历史版本需单独授权（下载历史版本）"
    # 当前推荐版本，按数据集下载策略
    policy = st.download_policy
    if policy == "forbidden":
        return False, "本数据集已关闭下载通道，仅可使用在线分析"
    if policy == "masked_only":
        if has_dataset_perm(db, d.id, user, "download.masked", version=ver):
            return True, "grant:download.masked"
        return False, "本数据集仅提供脱敏数据下载，需单独授权"
    if policy == "approval":
        if has_dataset_perm(db, d.id, user, "download.current", version=ver):
            return True, "grant:download.current"
        if _has_approved_download(db, d.id, user.id, ver):
            return True, "download_request"
        return False, "需提交下载申请并经管理员批准"
    if policy in ("member", "public"):
        return True, f"policy:{policy}"
    return False, "无下载权限"


@router.get("/datasets/{slug}/versions/{vid}/download")
def download(slug: str, vid: int, file: str = "data", db: Session = Depends(get_db),
             user: User = Depends(get_current_user)):
    d = _get_ds(db, slug)
    v = db.get(DataVersion, vid)
    if not v or v.dataset_id != d.id:
        raise HTTPException(404, "版本不存在")
    if file == "codebook":
        # 公开 codebook：成员可下；非成员仅当设置为 codebook_public 时可下
        st = get_settings(db, d.id)
        if not is_dataset_member(db, d.id, user) and not st.codebook_public:
            raise HTTPException(403, "该数据集 codebook 未公开，请先加入")
        source = "codebook"
        key = v.codebook_file_path
    else:
        ok, source = check_download(db, d, v, user)
        if not ok:
            raise HTTPException(403, source)
        key = v.data_file_path
    if not key:
        raise HTTPException(404, "文件不存在")
    # 四节 原则六 / 七节 4：留痕（含权限来源）
    db.add(DownloadLog(user_id=user.id, dataset_id=d.id, version_id=v.id, file_type=file,
                       file_name=key.split("/")[-1], downloaded_at=datetime.utcnow()))
    write_audit(db, user.id, "download", "dataset", d.id,
                {"version": v.version_id, "file": file, "source": source})
    db.commit()
    return StreamingResponse(storage.open(key), media_type="application/octet-stream",
                             headers={"Content-Disposition":
                                      f'attachment; filename="{key.split("/")[-1]}"'})


# -------- members / join --------
@router.get("/datasets/{slug}/members")
def members(slug: str, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    d = _get_ds(db, slug)
    if not is_dataset_member(db, d.id, user):
        raise HTTPException(403, "需为数据集成员")
    admin = is_dataset_admin(db, d.id, user)
    ms = db.query(DatasetMember).filter_by(dataset_id=d.id).all()

    def uname(uid):
        u = db.get(User, uid)
        return u.display_name if u else ""

    def perms_of(uid):
        gs = active_grants(db, d.id, uid)
        return sorted({g.perm for g in gs})

    lead_id = dataset_lead_id(db, d.id)
    out_members = [{"user_id": m.user_id, "name": uname(m.user_id), "ds_role": m.ds_role,
                    "is_admin": m.ds_role in DS_ADMIN_ROLES,
                    "is_lead": m.user_id == lead_id,
                    "perms": perms_of(m.user_id)} for m in ms]
    result = {"members": out_members, "is_lead": lead_id == user.id, "lead_id": lead_id}
    # 仅管理员可见待办：加入申请 + 下载申请
    if admin:
        reqs = db.query(JoinRequest).filter_by(dataset_id=d.id, status="pending").all()
        result["requests"] = [{"id": r.id, "user_id": r.user_id, "name": uname(r.user_id),
                               "status": r.status, "message": r.message} for r in reqs]
        dls = db.query(DownloadRequest).filter_by(dataset_id=d.id, status="pending").all()
        result["download_requests"] = [
            {"id": r.id, "user_id": r.user_id, "name": uname(r.user_id),
             "purpose": r.purpose, "scope_version": r.scope_version,
             "planned_until": r.planned_until, "share_with_others": r.share_with_others,
             "agree_charter": r.agree_charter} for r in dls]
        result["grant_catalog"] = [{"perm": p, "label": PERM_LABELS_ZH[p]}
                                   for p in GRANTABLE_PERMS]
    return result


# -------- 单独授权（六节：可带到期规则）--------
@router.post("/datasets/{slug}/members/{uid}/grant")
def grant(slug: str, uid: int, body: dict, db: Session = Depends(get_db),
          user: User = Depends(get_current_user)):
    """授予一项单独授权。body: {perm, scope_type, valid_to?, scope_version?, project_note?}。"""
    d = _get_ds(db, slug)
    if not is_dataset_admin(db, d.id, user):
        raise HTTPException(403, "仅数据集管理员可授权")
    m = dataset_membership(db, d.id, uid)
    if not m:
        raise HTTPException(404, "该用户不是成员")
    perm = body.get("perm")
    if perm not in GRANTABLE_PERMS:
        raise HTTPException(400, "未知的授权项")
    scope_type = body.get("scope_type", "permanent")
    valid_to = None
    if scope_type == "until_date" and body.get("valid_to"):
        try:
            valid_to = datetime.fromisoformat(body["valid_to"][:19])
        except ValueError:
            raise HTTPException(400, "有效期格式应为 ISO 日期")
    # 同一权限先撤旧再发新，避免叠加
    db.query(DatasetGrant).filter_by(dataset_id=d.id, user_id=uid, perm=perm,
                                     revoked=False).update({"revoked": True})
    g = DatasetGrant(dataset_id=d.id, user_id=uid, perm=perm, scope_type=scope_type,
                     valid_to=valid_to, scope_version=body.get("scope_version"),
                     project_note=body.get("project_note"), granted_by=user.id,
                     granted_at=datetime.utcnow())
    db.add(g)
    write_audit(db, user.id, "dataset.grant", "dataset", d.id,
                {"user": uid, "perm": perm, "scope": scope_type})
    db.commit()
    return {"ok": True, "grant_id": g.id}


@router.post("/datasets/{slug}/members/{uid}/revoke")
def revoke_grant(slug: str, uid: int, perm: str, db: Session = Depends(get_db),
                 user: User = Depends(get_current_user)):
    d = _get_ds(db, slug)
    if not is_dataset_admin(db, d.id, user):
        raise HTTPException(403, "仅数据集管理员可撤销授权")
    n = db.query(DatasetGrant).filter_by(dataset_id=d.id, user_id=uid, perm=perm,
                                         revoked=False).update({"revoked": True})
    write_audit(db, user.id, "dataset.revoke", "dataset", d.id, {"user": uid, "perm": perm})
    db.commit()
    return {"ok": True, "revoked": n}


# -------- 数据集管理员的增/删/交接：仅「总管理员」可操作 --------
@router.post("/datasets/{slug}/admins/{uid}")
def add_dataset_admin(slug: str, uid: int, db: Session = Depends(get_db),
                      user: User = Depends(get_current_user)):
    d = _get_ds(db, slug)
    if not is_dataset_lead(db, d.id, user):
        raise HTTPException(403, "仅数据集总管理员可设置管理员")
    m = dataset_membership(db, d.id, uid)
    if not m:
        raise HTTPException(404, "该用户不是成员，请先审批其加入")
    if m.ds_role in DS_LEAD_ROLES:
        raise HTTPException(400, "该用户已是总管理员")
    m.ds_role = "admin"
    write_audit(db, user.id, "dataset.admin.add", "dataset", d.id, {"user": uid})
    db.commit()
    return {"ok": True}


@router.delete("/datasets/{slug}/admins/{uid}")
def remove_dataset_admin(slug: str, uid: int, db: Session = Depends(get_db),
                         user: User = Depends(get_current_user)):
    d = _get_ds(db, slug)
    if not is_dataset_lead(db, d.id, user):
        raise HTTPException(403, "仅数据集总管理员可取消管理员")
    m = dataset_membership(db, d.id, uid)
    if not m or m.ds_role not in DS_ADMIN_ROLES:
        raise HTTPException(404, "该用户不是管理员")
    if m.ds_role in DS_LEAD_ROLES:
        raise HTTPException(400, "不能取消总管理员本人；请先把总管理员转让给他人")
    m.ds_role = "member"
    write_audit(db, user.id, "dataset.admin.remove", "dataset", d.id, {"user": uid})
    db.commit()
    return {"ok": True}


@router.post("/datasets/{slug}/transfer-lead/{uid}")
def transfer_dataset_lead(slug: str, uid: int, db: Session = Depends(get_db),
                          user: User = Depends(get_current_user)):
    """把「数据集总管理员」转让给另一名成员；原总管理员降为普通管理员。"""
    d = _get_ds(db, slug)
    if not is_dataset_lead(db, d.id, user):
        raise HTTPException(403, "仅数据集总管理员可转让")
    if uid == user.id:
        raise HTTPException(400, "不能转让给自己")
    target = dataset_membership(db, d.id, uid)
    if not target:
        raise HTTPException(404, "该用户不是成员，请先审批其加入")
    me = dataset_membership(db, d.id, user.id)
    target.ds_role = "owner"
    me.ds_role = "admin"
    write_audit(db, user.id, "dataset.lead.transfer", "dataset", d.id, {"to": uid})
    db.commit()
    return {"ok": True}


# -------- 数据集设置：下载策略 / 历史版本 / 在线分析 / 关闭（五、七、十节）--------
@router.patch("/datasets/{slug}/settings")
def update_settings(slug: str, body: dict, db: Session = Depends(get_db),
                    user: User = Depends(get_current_user)):
    d = _get_ds(db, slug)
    if not is_dataset_admin(db, d.id, user):
        raise HTTPException(403, "仅数据集管理员可设置")
    st = get_settings(db, d.id)
    if "download_policy" in body:
        if body["download_policy"] not in DOWNLOAD_POLICIES:
            raise HTTPException(400, "未知的下载策略")
        st.download_policy = body["download_policy"]
    for k in ("history_visible", "history_downloadable", "analysis_open",
              "codebook_public", "dashboard_public"):
        if k in body:
            setattr(st, k, bool(body[k]))
    write_audit(db, user.id, "dataset.settings", "dataset", d.id, body)
    db.commit()
    return {"ok": True}


@router.post("/datasets/{slug}/close")
def close_dataset(slug: str, closed: bool = True, db: Session = Depends(get_db),
                  user: User = Depends(get_current_user)):
    """关闭/重开数据集（十节 4）。彻底删除不作为普通页面按钮。"""
    d = _get_ds(db, slug)
    if not is_dataset_admin(db, d.id, user):
        raise HTTPException(403, "仅数据集管理员可关闭数据集")
    st = get_settings(db, d.id)
    st.is_closed = bool(closed)
    write_audit(db, user.id, "dataset.close" if closed else "dataset.reopen",
                "dataset", d.id)
    db.commit()
    return {"ok": True, "is_closed": st.is_closed}


# -------- 数据处理配置：唯一ID + 脱敏规则 --------
def _data_config(db, dataset_id) -> DatasetDataConfig:
    c = db.get(DatasetDataConfig, dataset_id)
    if not c:
        c = DatasetDataConfig(dataset_id=dataset_id)
        db.add(c); db.flush()
    return c


@router.get("/datasets/{slug}/data-config")
def get_data_config(slug: str, db: Session = Depends(get_db),
                    user: User = Depends(get_current_user)):
    d = _get_ds(db, slug)
    c = _data_config(db, d.id)
    rules = {r.var_name: {"mask_action": r.mask_action, "bucket_size": r.bucket_size}
             for r in db.query(VariableMaskRule).filter_by(dataset_id=d.id).all()}
    vs = db.query(Variable).filter_by(dataset_id=d.id, enabled=True).all()
    return {"unique_id_var": c.unique_id_var, "script_only": c.script_only,
            "variables": [{"var_name": v.var_name, "label_zh": v.label_zh,
                           "mask_action": rules.get(v.var_name, {}).get("mask_action", "keep"),
                           "bucket_size": rules.get(v.var_name, {}).get("bucket_size")}
                          for v in vs],
            "mask_actions": MASK_ACTIONS}


@router.put("/datasets/{slug}/data-config")
def set_data_config(slug: str, body: dict, db: Session = Depends(get_db),
                    user: User = Depends(get_current_user)):
    """设置唯一ID变量、是否仅脚本模式、以及各变量脱敏动作。"""
    d = _get_ds(db, slug)
    if not is_dataset_admin(db, d.id, user):
        raise HTTPException(403, "仅数据集管理员可配置")
    c = _data_config(db, d.id)
    if "unique_id_var" in body:
        c.unique_id_var = body["unique_id_var"] or None
    if "script_only" in body:
        c.script_only = bool(body["script_only"])
    # 脱敏规则（可选整表替换）
    if "rules" in body and isinstance(body["rules"], list):
        db.query(VariableMaskRule).filter_by(dataset_id=d.id).delete()
        for r in body["rules"]:
            act = r.get("mask_action", "keep")
            if act not in MASK_ACTIONS:
                continue
            db.add(VariableMaskRule(dataset_id=d.id, var_name=r["var_name"],
                                    mask_action=act, bucket_size=r.get("bucket_size")))
    write_audit(db, user.id, "dataset.data_config", "dataset", d.id)
    db.commit()
    return {"ok": True}


@router.post("/datasets/{slug}/versions/{vid}/desensitize")
def desensitize_version(slug: str, vid: int, new_version_id: str = Form(...),
                        changelog_zh: str = Form(""), db: Session = Depends(get_db),
                        user: User = Depends(get_current_user)):
    """按脱敏规则，从某原始版本一键生成脱敏版本。服务器可处理则直接建版；
    否则返回脚本供本地运行后再上传。"""
    from ..services.data_ops import desensitize
    d = _get_ds(db, slug)
    if not is_dataset_admin(db, d.id, user):
        raise HTTPException(403, "仅数据集管理员可生成脱敏版")
    src = db.get(DataVersion, vid)
    if not src or src.dataset_id != d.id:
        raise HTTPException(404, "源版本不存在")
    if not src.data_file_path:
        raise HTTPException(400, "源版本没有数据文件")
    if db.query(DataVersion).filter_by(dataset_id=d.id, version_id=new_version_id).first():
        raise HTTPException(400, f"版本 {new_version_id} 已存在")
    c = _data_config(db, d.id)
    rules = [{"var_name": r.var_name, "mask_action": r.mask_action,
              "bucket_size": r.bucket_size}
             for r in db.query(VariableMaskRule).filter_by(dataset_id=d.id).all()]
    new_bytes, source, script = desensitize(src.data_file_path, rules,
                                            c.unique_id_var, c.script_only)
    if new_bytes is None:
        # 回退：返回脚本，不建版
        write_audit(db, user.id, "dataset.desensitize.script", "dataset", d.id)
        return {"generated": "script", "script": script,
                "note": "数据过大或设为仅脚本模式：请在本地运行该脚本生成脱敏数据后，"
                        "用「发布新版本·脱敏」上传。"}
    key = f"versions/{d.slug}/{new_version_id}/masked.dta"
    storage.save(key, io.BytesIO(new_bytes))
    # 脱敏版参与迭代：取代当前推荐版
    db.query(DataVersion).filter_by(dataset_id=d.id, is_current=True).update(
        {"is_current": False, "valid_to": datetime.utcnow()})
    v = DataVersion(dataset_id=d.id, version_id=new_version_id,
                    based_on_version=src.version_id, release_date=datetime.utcnow(),
                    data_file_path=key, changelog_zh=changelog_zh or f"由 {src.version_id} 脱敏生成",
                    created_by=user.id, is_current=True, valid_from=datetime.utcnow())
    db.add(v); db.flush()
    db.add(VersionExtra(version_id=v.id, data_kind="masked",
                        masked_source_version=src.version_id, generated=source))
    d.current_version_id = v.id
    write_audit(db, user.id, "dataset.desensitize.server", "dataset", d.id,
                {"from": src.version_id, "to": new_version_id})
    db.commit()
    return {"generated": "server", "id": v.id, "version_id": new_version_id}


# -------- 下载申请与审批（七节 3）--------
@router.post("/datasets/{slug}/download-requests")
def create_download_request(slug: str, body: dict, db: Session = Depends(get_db),
                            user: User = Depends(get_current_user)):
    d = _get_ds(db, slug)
    if not is_dataset_member(db, d.id, user):
        raise HTTPException(403, "需先加入数据集才能申请下载")
    if not body.get("agree_charter"):
        raise HTTPException(400, "需同意数据使用公约")
    if not (body.get("purpose") or "").strip():
        raise HTTPException(400, "请填写研究用途")
    r = DownloadRequest(dataset_id=d.id, user_id=user.id, purpose=body.get("purpose"),
                        scope_version=body.get("scope_version"),
                        planned_until=body.get("planned_until"),
                        share_with_others=bool(body.get("share_with_others")),
                        agree_charter=True, status="pending",
                        created_at=datetime.utcnow())
    db.add(r); db.flush()
    write_audit(db, user.id, "download.request", "dataset", d.id, {"req": r.id})
    db.commit()
    return {"id": r.id, "status": "pending"}


@router.post("/download-requests/{rid}/decide")
def decide_download_request(rid: int, approve: bool, valid_to: str = "",
                            scope_version: str = "", db: Session = Depends(get_db),
                            user: User = Depends(get_current_user)):
    r = db.get(DownloadRequest, rid)
    if not r or r.status != "pending":
        raise HTTPException(404, "申请不存在或已处理")
    if not is_dataset_admin(db, r.dataset_id, user):
        raise HTTPException(403, "仅数据集管理员可审批下载")
    r.status = "approved" if approve else "rejected"
    r.decided_by = user.id; r.decided_at = datetime.utcnow()
    if scope_version:
        r.scope_version = scope_version
    if approve and valid_to:
        try:
            r.valid_to = datetime.fromisoformat(valid_to[:19])
        except ValueError:
            raise HTTPException(400, "有效期格式应为 ISO 日期")
    write_audit(db, user.id, "download.request.decide", "dataset", r.dataset_id,
                {"req": rid, "approve": approve})
    db.commit()
    return {"ok": True, "status": r.status}


@router.post("/download-requests/{rid}/revoke")
def revoke_download_request(rid: int, db: Session = Depends(get_db),
                            user: User = Depends(get_current_user)):
    r = db.get(DownloadRequest, rid)
    if not r:
        raise HTTPException(404, "申请不存在")
    if not is_dataset_admin(db, r.dataset_id, user):
        raise HTTPException(403, "仅数据集管理员可撤销")
    r.status = "revoked"
    write_audit(db, user.id, "download.request.revoke", "dataset", r.dataset_id, {"req": rid})
    db.commit()
    return {"ok": True}


# -------- 版本候选文件（九节 1：获授权成员上传，管理员发布）--------
@router.get("/datasets/{slug}/candidates")
def list_candidates(slug: str, db: Session = Depends(get_db),
                    user: User = Depends(get_current_user)):
    d = _get_ds(db, slug)
    if not is_dataset_member(db, d.id, user):
        raise HTTPException(403, "需为数据集成员")
    cs = db.query(VersionCandidate).filter_by(dataset_id=d.id).order_by(
        VersionCandidate.id.desc()).all()
    return [{"id": c.id, "file_name": c.file_name, "note": c.note, "status": c.status,
             "uploaded_by": c.uploaded_by,
             "created_at": str(c.created_at) if c.created_at else None} for c in cs]


@router.post("/datasets/{slug}/candidates")
def upload_candidate(slug: str, note: str = Form(""), file: UploadFile = File(...),
                     db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    d = _get_ds(db, slug)
    if get_settings(db, d.id).is_closed:
        raise HTTPException(400, "数据集已关闭，不再接受候选文件")
    if not has_dataset_perm(db, d.id, user, "upload.candidate"):
        raise HTTPException(403, "需获得「上传版本候选文件」授权")
    key = f"candidates/{d.slug}/{datetime.utcnow():%Y%m%d%H%M%S}_{file.filename}"
    storage.save(key, file.file)
    c = VersionCandidate(dataset_id=d.id, uploaded_by=user.id, file_path=key,
                         file_name=file.filename, note=note, status="pending",
                         created_at=datetime.utcnow())
    db.add(c); db.flush()
    write_audit(db, user.id, "version.candidate.upload", "dataset", d.id, {"cand": c.id})
    db.commit()
    return {"id": c.id, "file_name": c.file_name}


# -------- codebook 草稿（六节：codebook.draft 授权）--------
@router.get("/datasets/{slug}/codebook-draft")
def get_codebook_draft(slug: str, db: Session = Depends(get_db),
                       user: User = Depends(get_current_user)):
    d = _get_ds(db, slug)
    if not is_dataset_member(db, d.id, user):
        raise HTTPException(403, "需为数据集成员")
    draft = db.query(CodebookDraft).filter_by(dataset_id=d.id).order_by(
        CodebookDraft.id.desc()).first()
    can_edit = has_dataset_perm(db, d.id, user, "codebook.draft")
    return {"body_zh": draft.body_zh if draft else "", "can_edit": can_edit,
            "updated_at": str(draft.updated_at) if draft and draft.updated_at else None}


@router.put("/datasets/{slug}/codebook-draft")
def save_codebook_draft(slug: str, body: dict, db: Session = Depends(get_db),
                        user: User = Depends(get_current_user)):
    d = _get_ds(db, slug)
    if not has_dataset_perm(db, d.id, user, "codebook.draft"):
        raise HTTPException(403, "需获得「编辑 codebook 草稿」授权")
    draft = db.query(CodebookDraft).filter_by(dataset_id=d.id).order_by(
        CodebookDraft.id.desc()).first()
    if not draft:
        draft = CodebookDraft(dataset_id=d.id)
        db.add(draft)
    draft.body_zh = body.get("body_zh", "")
    draft.updated_by = user.id; draft.updated_at = datetime.utcnow()
    write_audit(db, user.id, "codebook.draft.save", "dataset", d.id)
    db.commit()
    return {"ok": True}


@router.post("/datasets/{slug}/join-requests")
def join_dataset(slug: str, message: str = "", db: Session = Depends(get_db),
                 user: User = Depends(get_current_user)):
    d = _get_ds(db, slug)
    if get_settings(db, d.id).is_closed:
        raise HTTPException(400, "该数据集已关闭，不再接受新成员申请")
    if dataset_membership(db, d.id, user.id):
        raise HTTPException(400, "已是成员")
    if db.query(JoinRequest).filter_by(dataset_id=d.id, user_id=user.id,
                                       status="pending").first():
        raise HTTPException(400, "已提交申请，等待审批")
    r = JoinRequest(dataset_id=d.id, user_id=user.id, message=message, status="pending")
    db.add(r); db.commit()
    return {"id": r.id}


@router.post("/join-requests/{rid}/decide")
def decide_join(rid: int, approve: bool, db: Session = Depends(get_db),
                user: User = Depends(get_current_user)):
    r = db.get(JoinRequest, rid)
    if not r:
        raise HTTPException(404, "申请不存在")
    if not is_dataset_admin(db, r.dataset_id, user):
        raise HTTPException(403, "仅发起人/管理员可审批")
    r.status = "approved" if approve else "rejected"
    r.decided_by = user.id; r.decided_at = datetime.utcnow()
    if approve:
        # 原则三：加入数据集 ≠ 获得下载权。成员默认无任何单独授权，需管理员另行授予。
        db.add(DatasetMember(dataset_id=r.dataset_id, user_id=r.user_id, ds_role="member",
                             granted_perms_json=[],
                             joined_at=datetime.utcnow(), approved_by=user.id))
    write_audit(db, user.id, "dataset.join.decide", "dataset", r.dataset_id,
                {"approve": approve, "applicant": r.user_id})
    db.commit()
    return {"ok": True, "status": r.status}


# -------- 数据看板（从派生汇总表出图，只读）--------
@router.get("/datasets/{slug}/dashboard")
def dashboard(slug: str, var: str, group: str = "", db: Session = Depends(get_db),
              user: User = Depends(get_current_user)):
    d = _get_ds(db, slug)
    # 公开看板：成员可看；非成员仅当设置为 dashboard_public 时可看
    if not is_dataset_member(db, d.id, user) and not get_settings(db, d.id).dashboard_public:
        raise HTTPException(403, "该数据集看板未公开，请先加入")
    q = db.query(DatasetSummary).filter_by(dataset_id=d.id, var_name=var)
    if group:
        q = q.filter_by(group_key=group)
    rows = q.all()
    return [{"bucket": r.bucket, "value": r.value, "group_key": r.group_key} for r in rows]


def _require_analysis(db, d, user):
    """在线分析需授权（原则三 / 速查表）。管理员或获 analysis.online 授权，或数据集开放。"""
    if not is_dataset_member(db, d.id, user):
        raise HTTPException(403, "需先加入数据集")
    if is_dataset_admin(db, d.id, user):
        return
    if get_settings(db, d.id).analysis_open:
        return
    if not has_dataset_perm(db, d.id, user, "analysis.online"):
        raise HTTPException(403, "在线分析需单独授权")


@router.post("/datasets/{slug}/analysis/generate")
def gen_analysis(slug: str, prompt: str, db: Session = Depends(get_db),
                 user: User = Depends(get_current_user)):
    """AI 只接收变量名/codebook/需求，返回分析代码；不接收原始数据行。"""
    d = _get_ds(db, slug)
    _require_analysis(db, d, user)
    vs = db.query(Variable).filter_by(dataset_id=d.id, enabled=True).all()
    var_list = ", ".join(f"{v.var_name}" + (f"（{v.label_zh}）" if v.label_zh else "") for v in vs)
    groups = sorted({r.group_key for r in
                     db.query(DatasetSummary).filter_by(dataset_id=d.id).all() if r.group_key})
    vars_in_summary = sorted({r.var_name for r in
                              db.query(DatasetSummary).filter_by(dataset_id=d.id).all()})
    # 关键：明确告诉 AI 沙箱里 df 的真实结构（派生汇总长表），避免用不存在的原始列名
    sys = (
        "你是数据分析助手，为一个只读沙箱写 pandas 代码。沙箱里已有一个名为 df 的 DataFrame，"
        "它不是原始个体数据，而是【派生汇总长表】，恰好有四列：\n"
        "  var_name(变量名) | group_key(分组，如性别/年份) | bucket(取值区间/类别) | value(该桶的计数或统计量)\n"
        "因此要按某变量分析时，应先 df[df['var_name']=='变量名']，再对 bucket/value 聚合，"
        "不要直接使用原始变量名作为列名。把最终结果赋值给变量 result。禁止文件/网络/写操作。")
    ctx = (f"该数据集包含的变量：{var_list}\n"
           f"汇总表里已有的 var_name：{', '.join(vars_in_summary) or '（暂无）'}\n"
           f"汇总表里已有的 group_key：{', '.join(groups) or '（无分组）'}\n"
           f"分析需求：{prompt}\n请只输出 pandas 代码：")
    code = ai_client.complete(ctx, sys)
    return {"code": code, "lang": "Python"}


@router.post("/datasets/{slug}/analysis/run")
def run_analysis(slug: str, code: str, db: Session = Depends(get_db),
                 user: User = Depends(get_current_user)):
    """只读沙箱执行；使用派生汇总，不加载敏感原始行。"""
    from ..services.sandbox import run_readonly, SandboxViolation
    d = _get_ds(db, slug)
    _require_analysis(db, d, user)
    rows = db.query(DatasetSummary).filter_by(dataset_id=d.id).all()
    records = [{"var_name": r.var_name, "group_key": r.group_key,
                "bucket": r.bucket, "value": r.value} for r in rows]
    try:
        return run_readonly(code, records)
    except SandboxViolation as e:
        raise HTTPException(400, str(e))


# -------- literature / publications --------
@router.get("/datasets/{slug}/literature")
def literature(slug: str, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    d = _get_ds(db, slug)
    topics = db.query(LitTopic).filter_by(dataset_id=d.id).all()
    refs = db.query(LitRef).filter_by(dataset_id=d.id).all()
    return {"topics": [{"id": t.id, "parent_id": t.parent_id, "title_zh": t.title_zh,
                        "ai_generated": t.ai_generated} for t in topics],
            "refs": [{"id": r.id, "title": r.title, "authors": r.authors,
                      "venue": r.venue, "year": r.year, "url": r.url,
                      "doi": getattr(r, "doi", None), "note_zh": r.note_zh,
                      "citation": _citation(r.title, r.authors, r.year, r.venue,
                                            getattr(r, "doi", None), r.url)}
                     for r in refs]}


@router.post("/datasets/{slug}/literature/refs")
def add_ref(slug: str, body: LitRefIn, db: Session = Depends(get_db),
            user: User = Depends(get_current_user)):
    d = _get_ds(db, slug)
    if not is_dataset_member(db, d.id, user):
        raise HTTPException(403, "需为数据集成员")
    if not (body.title and body.authors and body.year and body.venue):
        raise HTTPException(400, "标题、作者、年份、刊物/出版社为必填")
    # 单条也做 AI 真实性核验：可疑且未确认真实 → 拦截并回传结论
    verdict = _ai_verify_refs([{"title": body.title, "authors": body.authors,
                                "year": body.year, "venue": body.venue, "doi": body.doi}])[0]
    if verdict["verdict"] == "suspect" and not body.confirm_real:
        return {"ok": False, "verdict": verdict,
                "detail": "AI 判定该文献可疑，请核对后勾选「确认真实文献」再提交"}
    data = body.model_dump(exclude={"confirm_real"})
    r = LitRef(dataset_id=d.id, added_by=user.id, **data)
    db.add(r); db.commit()
    return {"ok": True, "id": r.id, "verdict": verdict}


@router.post("/datasets/{slug}/literature/parse-citation")
def parse_citation(slug: str, body: CitationTextIn, db: Session = Depends(get_db),
                   user: User = Depends(get_current_user)):
    """输入一段引文文本，自动解析为「标题/作者/年份/刊物/DOI/URL」字段（AI 优先，回退正则）。"""
    d = _get_ds(db, slug)
    if not is_dataset_member(db, d.id, user):
        raise HTTPException(403, "需为数据集成员")
    text = (body.text or "").strip()
    if not text:
        raise HTTPException(400, "请粘贴一段引文")
    import re
    fields = {"title": "", "authors": "", "year": None, "venue": "", "doi": None, "url": None}
    # 通用抽取：年份、DOI、URL
    ym = re.search(r"(19|20)\d{2}", text)
    if ym:
        fields["year"] = int(ym.group(0))
    dm = re.search(r"10\.\d{4,9}/[-._;()/:A-Za-z0-9]+", text)
    if dm:
        fields["doi"] = dm.group(0)
    um = re.search(r"https?://[^\s]+", text)
    if um:
        fields["url"] = um.group(0)
    if ai_client.enabled():
        sys = ("你是文献解析助手。把用户给的一条引文解析为 JSON，"
               '字段：title, authors, year(数字), venue(刊物或出版社), doi, url。'
               "只输出 JSON，无法确定的留空字符串或 null。")
        import json
        resp = ai_client.complete(text, sys)
        try:
            js = json.loads(resp[resp.find("{"): resp.rfind("}") + 1])
            for k in fields:
                if js.get(k) not in (None, ""):
                    fields[k] = js.get(k)
            if fields.get("year"):
                try: fields["year"] = int(str(fields["year"])[:4])
                except Exception: pass
        except Exception:
            pass
    # 正则兜底 title/authors：以句点切分，含年份的前段常为"作者(年)."，其后为标题
    if not fields["title"]:
        segs = [s.strip() for s in re.split(r"[.。]", text) if s.strip()]
        if segs:
            if len(segs) >= 2 and ym:
                fields["authors"] = fields["authors"] or segs[0]
                fields["title"] = segs[1]
                if len(segs) >= 3:
                    fields["venue"] = fields["venue"] or segs[2]
            else:
                fields["title"] = segs[0]
    fields["citation"] = _citation(fields["title"], fields["authors"], fields["year"],
                                   fields["venue"], fields["doi"], fields["url"])
    return {"fields": fields, "ai_model": ai_client.provider}


@router.get("/datasets/{slug}/literature/export")
def export_lit(slug: str, db: Session = Depends(get_db),
               user: User = Depends(get_current_user)):
    """一键导出全部文献为 Excel（列与上传模板一致）。"""
    from openpyxl import Workbook
    d = _get_ds(db, slug)
    wb = Workbook(); ws = wb.active; ws.title = "文献"
    ws.append(LIT_COLS)
    for r in db.query(LitRef).filter_by(dataset_id=d.id).all():
        ws.append([r.title, r.authors, r.year, r.venue,
                   getattr(r, "doi", None) or "", r.url or ""])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{d.slug}_literature.xlsx"'})


# ================= 文献批量导入 / 引用格式 / AI 真实性核验 =================
LIT_COLS = ["标题", "作者", "年份", "刊物/出版社", "DOI", "链接URL"]
LIT_REQUIRED = ["标题", "作者", "年份", "刊物/出版社"]


def _citation(title, authors, year, venue, doi=None, url=None) -> str:
    """生成一条参考文献引用格式（作者(年份). 标题. 刊物. DOI/URL）。"""
    parts = []
    if authors:
        parts.append(f"{authors}")
    if year:
        parts.append(f"({year}).")
    elif authors:
        parts[-1] = parts[-1] + "."
    if title:
        parts.append(f"{title}.")
    if venue:
        parts.append(f"{venue}.")
    if doi:
        parts.append(f"https://doi.org/{doi}" if not str(doi).startswith("http") else str(doi))
    elif url:
        parts.append(str(url))
    return " ".join(parts).strip()


@router.get("/datasets/{slug}/lit-template")
def lit_template(slug: str, db: Session = Depends(get_db),
                 user: User = Depends(get_current_user)):
    """下载批量文献 Excel 模板（列头 + 填写说明）。"""
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    d = _get_ds(db, slug)
    wb = Workbook()
    ws = wb.active; ws.title = "文献"
    ws.append(LIT_COLS)
    notes = {
        "标题": "文献标题（必填）。",
        "作者": "作者，多位可用逗号或分号分隔（必填）。",
        "年份": "发表年份，如 2021（必填）。",
        "刊物/出版社": "期刊名或出版社（必填）。",
        "DOI": "数字对象唯一标识（可选）。",
        "链接URL": "可点击跳转的链接（可选）。",
    }
    for i, col in enumerate(LIT_COLS, start=1):
        ws.cell(row=1, column=i).comment = Comment(notes[col], "系统")
    ws2 = wb.create_sheet("填写说明")
    for line in [
        "批量文献填写说明：",
        "1. 每一行是一条参考文献；导入后系统按行拆分预览。",
        "2. 标题、作者、年份、刊物/出版社为必填；DOI、链接URL 可选。",
        "3. 导入后可一键 AI 核验是否为真实文献；被判可疑的会标出，",
        "   你核对后仍可勾选「确认真实文献」强制上传（以防 AI 误判）。",
        "4. 支持 .xlsx / .csv，列顺序：" + "、".join(LIT_COLS) + "。",
    ]:
        ws2.append([line])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{d.slug}_lit_template.xlsx"'})


def _parse_lit(raw: bytes, filename: str) -> list[dict]:
    name = (filename or "").lower()
    rows = []
    if name.endswith(".csv"):
        import csv
        for r in csv.DictReader(io.StringIO(raw.decode("utf-8-sig", errors="replace"))):
            if any((v or "").strip() for v in r.values()):
                rows.append(r)
    else:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb["文献"] if "文献" in wb.sheetnames else wb.worksheets[0]
        header = None
        for row in ws.iter_rows(values_only=True):
            if header is None:
                header = [str(c).strip() if c is not None else "" for c in row]; continue
            if row is None or all(c is None for c in row):
                continue
            rows.append({header[i]: row[i] for i in range(min(len(header), len(row)))})
    out = []
    for r in rows:
        y = r.get("年份")
        try:
            y = int(str(y).strip()[:4]) if y not in (None, "") else None
        except Exception:
            y = None
        item = {"title": str(r.get("标题", "") or "").strip(),
                "authors": str(r.get("作者", "") or "").strip(),
                "year": y, "venue": str(r.get("刊物/出版社", "") or "").strip(),
                "doi": str(r.get("DOI", "") or "").strip() or None,
                "url": str(r.get("链接URL", "") or "").strip() or None}
        missing = [c for c in LIT_REQUIRED
                   if not (item["title"] if c == "标题" else item["authors"] if c == "作者"
                           else item["year"] if c == "年份" else item["venue"])]
        item["missing"] = missing
        item["citation"] = _citation(item["title"], item["authors"], item["year"],
                                      item["venue"], item["doi"], item["url"])
        out.append(item)
    return out


@router.post("/datasets/{slug}/literature/parse")
def parse_lit(slug: str, file: UploadFile = File(...), db: Session = Depends(get_db),
              user: User = Depends(get_current_user)):
    """解析批量文献文件（不落库），返回逐条预览 + 必填校验 + 引用格式。"""
    d = _get_ds(db, slug)
    if not is_dataset_member(db, d.id, user):
        raise HTTPException(403, "需为数据集成员")
    rows = _parse_lit(file.file.read(), file.filename)
    if not rows:
        raise HTTPException(400, "未解析到有效文献行，请用模板填写")
    return {"rows": rows}


def _ai_verify_refs(refs: list[dict]) -> list[dict]:
    """用 AI 判断每条文献是否可能真实存在。返回 [{index, verdict, reason}]。
    verdict: real | suspect | unknown（AI 未启用或解析失败时为 unknown，放行）。"""
    if not ai_client.enabled():
        return [{"index": i, "verdict": "unknown",
                 "reason": "AI 未启用，未做真实性核验"} for i in range(len(refs))]
    listing = "\n".join(
        f"{i+1}. 标题《{r.get('title','')}》；作者 {r.get('authors','')}；"
        f"年份 {r.get('year','')}；刊物/出版社 {r.get('venue','')}；"
        f"DOI {r.get('doi') or '无'}" for i, r in enumerate(refs))
    sys = ("你是学术文献核验助手。判断每条参考文献是否可能是真实存在的文献。"
           "对每条只输出一行，格式：序号|real 或 suspect|简短理由。"
           "拿不准偏向 suspect。只输出这些行，不要多余内容。")
    resp = ai_client.complete(listing, sys)
    verdicts = {}
    for line in (resp or "").splitlines():
        parts = line.split("|")
        if len(parts) >= 2:
            num = "".join(c for c in parts[0] if c.isdigit())
            if num:
                v = parts[1].strip().lower()
                verdict = "real" if "real" in v else ("suspect" if "suspect" in v else "unknown")
                verdicts[int(num) - 1] = {"verdict": verdict,
                                          "reason": (parts[2].strip() if len(parts) > 2 else "")}
    return [{"index": i,
             "verdict": verdicts.get(i, {}).get("verdict", "unknown"),
             "reason": verdicts.get(i, {}).get("reason", "未能解析 AI 结果，请人工确认")}
            for i in range(len(refs))]


@router.post("/datasets/{slug}/literature/ai-verify")
def ai_verify_lit(slug: str, body: LitBatchIn, db: Session = Depends(get_db),
                  user: User = Depends(get_current_user)):
    """对一批文献做 AI 真实性核验，返回逐条结论（不落库）。"""
    d = _get_ds(db, slug)
    if not is_dataset_member(db, d.id, user):
        raise HTTPException(403, "需为数据集成员")
    refs = [r.model_dump() for r in body.refs]
    return {"results": _ai_verify_refs(refs), "ai_model": ai_client.provider}


@router.post("/datasets/{slug}/literature/batch")
def batch_commit_lit(slug: str, body: LitBatchIn, db: Session = Depends(get_db),
                     user: User = Depends(get_current_user)):
    """批量落库：必填校验；AI 判为可疑且未勾选「确认真实」的条目会被拒绝并返回。"""
    d = _get_ds(db, slug)
    if not is_dataset_member(db, d.id, user):
        raise HTTPException(403, "需为数据集成员")
    refs = [r.model_dump() for r in body.refs]
    if not refs:
        raise HTTPException(400, "没有可导入的文献")
    # 必填校验
    for i, r in enumerate(refs):
        if not (r.get("title") and r.get("authors") and r.get("year") and r.get("venue")):
            raise HTTPException(400, f"第 {i+1} 条缺少必填项（标题/作者/年份/刊物）")
    # AI 核验：可疑且未确认 → 拦截
    verdicts = _ai_verify_refs(refs)
    blocked = []
    for v in verdicts:
        i = v["index"]
        if v["verdict"] == "suspect" and not refs[i].get("confirm_real"):
            blocked.append({"index": i, "title": refs[i].get("title"), "reason": v["reason"]})
    if blocked:
        return {"ok": False, "blocked": blocked, "verdicts": verdicts,
                "detail": "以下文献被 AI 判为可疑，请核对后勾选「确认真实文献」再导入"}
    created = 0
    for r in refs:
        db.add(LitRef(dataset_id=d.id, added_by=user.id, title=r["title"],
                      authors=r.get("authors"), venue=r.get("venue"), year=r.get("year"),
                      doi=r.get("doi"), url=r.get("url")))
        created += 1
    db.commit()
    return {"ok": True, "created": created}


# ================= AI 文献/用途总结（元数据-only，安全）=================
@router.post("/datasets/{slug}/literature/ai-summarize")
def ai_summarize_literature(slug: str, body: AiSummaryIn | None = None,
                            db: Session = Depends(get_db),
                            user: User = Depends(get_current_user)):
    """基于「所有文献标题」做定制化综述。支持用户自定义提示词（如"按方法分类""指出研究空白"）。"""
    d = _get_ds(db, slug)
    if not is_dataset_member(db, d.id, user):
        raise HTTPException(403, "需为数据集成员")
    refs = db.query(LitRef).filter_by(dataset_id=d.id).all()
    if not refs:
        return {"summary": "该数据集暂无文献，添加文献后即可生成综述。", "ai_model": ai_client.provider}
    titles = "\n".join(f"- {r.title}（{r.authors or ''}，{r.venue or ''}，{r.year or ''}）" for r in refs)
    user_prompt = (body.prompt if body else "") or ""
    ctx = (f"数据集：{d.name_zh}。共有 {len(refs)} 篇文献，标题清单如下：\n{titles}\n\n"
           + (f"用户额外要求：{user_prompt}\n" if user_prompt.strip() else ""))
    sys = ("你是文献综述助理。仅依据上面给出的文献标题信息，用中文写一段结构化的文献综述："
           "先概述整体主题分布，再归纳 3-6 个研究方向/主题（每个方向列出相关文献），"
           "最后指出可能的研究空白。若用户有额外要求，请优先满足。"
           "不要编造标题里没有的文献。")
    out = ai_client.complete(ctx, sys, strong=True)
    return {"summary": out, "ai_model": ai_client.provider, "count": len(refs)}


# ================= AI 勘误提示（悬浮助手；只提示不改数，保数据安全）=================
# 预设勘误方向（会随平台勘误记录增长，由 patterns 模式从历史中总结补充）
BASE_CHECK_DIRECTIONS = [
    "缺失值：是否有关键变量存在异常比例的空缺或占位符（如 -99、NA、空字符串）。",
    "取值范围：数值变量是否有超出合理范围的极端值/离群值（如年龄为负、比例>1）。",
    "唯一性：唯一ID是否真的唯一，是否存在重复记录或同一主体多条冲突记录。",
    "一致性：同一实体在不同年份/来源的属性是否前后矛盾（如性别、出生年变化）。",
    "编码口径：分类变量的编码是否统一（同义不同码、口径中途变更）。",
    "时间连续性：面板数据是否有断档、时间倒挂或重复年份。",
    "单位与量纲：金额/比率是否存在单位不一致（万元 vs 元、百分比 vs 小数）。",
    "文本规范：名称/机构等字符串是否有多余空格、全半角混用、别名未合并。",
]


@router.get("/datasets/{slug}/check-directions")
def check_directions(slug: str, db: Session = Depends(get_db),
                     user: User = Depends(get_current_user)):
    d = _get_ds(db, slug)
    return {"directions": BASE_CHECK_DIRECTIONS,
            "unique_id_var": _uid_var_name(db, d.id)}


def _uid_var_name(db, ds_id):
    cfg = db.get(DatasetDataConfig, ds_id)
    return cfg.unique_id_var if cfg else None


@router.post("/datasets/{slug}/ai-hint")
def ai_hint(slug: str, body: AiHintIn, db: Session = Depends(get_db),
            user: User = Depends(get_current_user)):
    """AI 勘误提示：只用元数据（变量清单、已有勘误理由），提示「哪些方面可能有问题、
    建议人工核查」，绝不读取原始数据、绝不自动改数。点击一次运行一次。

    mode=check：结合预设方向 + 变量清单 + 人工提示词，给出需人工核查的维度清单。
    mode=patterns：从本数据集/平台已有勘误记录中总结高频错误类型与注意事项。
    """
    from ..models.curation import BugItem
    d = _get_ds(db, slug)
    if not is_dataset_member(db, d.id, user):
        raise HTTPException(403, "需为数据集成员")
    vs = db.query(Variable).filter_by(dataset_id=d.id).all()
    var_list = ", ".join(f"{v.var_name}({v.label_zh})" if v.label_zh else v.var_name for v in vs)
    prompt = (body.prompt or "").strip()

    if body.mode == "patterns":
        # 汇总已有勘误理由（先本数据集，样本不足则并入全平台），只用文本、不含原始数据
        reasons = [b.description_zh for b in db.query(Bug).filter_by(dataset_id=d.id).all()
                   if b.description_zh]
        reasons += [it.reason for it in db.query(BugItem).filter_by(dataset_id=d.id).all()
                    if it.reason]
        if len(reasons) < 5:
            reasons += [b.description_zh for b in db.query(Bug).limit(200).all()
                        if b.description_zh][:50]
        corpus = "\n".join(f"- {r}" for r in reasons[:120]) or "（暂无历史勘误记录）"
        sys = ("你是数据质量分析助手。下面是本平台已提交的勘误理由清单，"
               "请归纳出现频率较高的错误类型，并按'高频→低频'给出注意事项清单，"
               "帮助后续勘误者重点关注。只输出归纳，不要编造。")
        out = ai_client.complete(corpus, sys)
        return {"hint": out, "mode": "patterns", "ai_model": ai_client.provider,
                "n_records": len(reasons)}

    # check 模式
    base = "\n".join(f"- {x}" for x in BASE_CHECK_DIRECTIONS)
    ctx = (f"数据集：{d.name_zh}。简介：{d.desc_zh or ''}。\n"
           f"变量清单：{var_list or '（未登记变量）'}。\n"
           f"唯一ID变量：{_uid_var_name(db, d.id) or '未设置'}。\n"
           f"通用检查方向：\n{base}\n"
           + (f"\n本次人工限定的方向/提示：{prompt}\n" if prompt else ""))
    sys = ("你是数据质量核查助手。重要原则：你看不到也不需要原始数据，"
           "只能基于变量清单与研究背景，提示『哪些方面/哪些变量可能存在问题、建议人工去核查』，"
           "绝不断言数据一定有错，绝不生成或修改数据。用中文输出分点清单，"
           "每点指出：可能的问题、涉及哪个/哪类变量、建议如何人工核查。"
           "若有人工限定方向，请围绕该方向展开。")
    out = ai_client.complete(ctx, sys)
    return {"hint": out, "mode": "check", "ai_model": ai_client.provider}
